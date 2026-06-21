import os
import glob
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

def clean_card(val):
    if pd.isna(val):
        return None
    s = str(val).strip()
    if s.endswith('.0'):
        s = s[:-2]
    s = s.replace(" ", "")
    if len(s) == 16 and s.isdigit():
        return s
    return None

def clean_account(val):
    if pd.isna(val):
        return None
    s = str(val).strip()
    if s.endswith('.0'):
        s = s[:-2]
    s = s.replace(" ", "").replace("-", "")
    if s.isdigit():
        # Excel's TEXT(..., "00000000000") pads to at least 11 digits
        return s.zfill(11)
    return None

def clean_dni(val):
    if pd.isna(val):
        return None
    s = str(val).strip()
    if s.endswith('.0'):
        s = s[:-2]
    s = s.replace(" ", "")
    
    # Validacion de RUC (11 digitos, empieza con 10, 15, 17 o 20)
    if len(s) == 11 and s.isdigit() and s[:2] in ["10", "15", "17", "20"]:
        return s
    
    # Passports (0X...)
    if s.upper().startswith("0X"):
        return s
    
    # Carne de extranjeria starting with 5 zeros (e.g. 00000...)
    if s.startswith("00000"):
        return s
    
    # Also handle CE- prefix containing 5 zeros
    temp_s = s
    if temp_s.upper().startswith("CE-"):
        temp_s = temp_s[3:]
    elif temp_s.upper().startswith("CE"):
        temp_s = temp_s[2:]
        
    if temp_s.startswith("00000"):
        return temp_s
        
    # Standard DNI
    if temp_s.isdigit() and len(temp_s) <= 8:
        return temp_s.zfill(8)
            
    return None

def clean_phone(val):
    if pd.isna(val):
        return None
    s = str(val).strip()
    if s.endswith('.0'):
        s = s[:-2]
    s = s.replace(" ", "")
    if len(s) == 9 and s.isdigit():
        if s.startswith("9"):
            if s not in ["999999999", "900000000", "111111111"]:
                return s
    return None

def to_str_no_sci(val):
    if pd.isna(val):
        return ""
    if isinstance(val, (int, float)):
        if isinstance(val, float):
            # Check if it's actually an integer-like float
            if val.is_integer():
                return str(int(val))
            # Format float to avoid scientific notation
            s = f"{val:.10f}".rstrip('0')
            if s.endswith('.'):
                s = s[:-1]
            return s
        else:
            return str(val)
    
    s = str(val).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s

def main():
    print("=== INICIANDO PROCESO DE LIMPIEZA DE BASE ===")
    
    # Find the file in LIMPIAR directory
    search_dir = "LIMPIAR"
    excel_files = glob.glob(os.path.join(search_dir, "*GESTION PENDIENTES*.xlsx")) + \
                  glob.glob(os.path.join(search_dir, "*gestion pendientes*.xlsx")) + \
                  glob.glob(os.path.join(search_dir, "*.xlsx"))
                  
    # Remove duplicates from glob
    excel_files = list(dict.fromkeys(excel_files))
    
    if not excel_files:
        print(f"Error: No se encontró ningún archivo Excel en la carpeta {search_dir}.")
        return
        
    # Prefer files with "GESTION PENDIENTES" in their name
    target_file = None
    for f in excel_files:
        if "GESTION PENDIENTES" in f.upper():
            target_file = f
            break
    if not target_file:
        target_file = excel_files[0]
        
    print(f"Archivo detectado: {target_file}")
    
    # Load workbook to list sheet names
    try:
        wb_check = openpyxl.load_workbook(target_file, read_only=True)
        sheets = wb_check.sheetnames
        print("\nHojas disponibles en el archivo:")
        for idx, sheet in enumerate(sheets, 1):
            print(f" {idx}. {sheet}")
    except Exception as e:
        print(f"Error al leer el archivo Excel: {e}")
        return
        
    # Let user input the sheet name
    sheet_name = input("\nIngrese el nombre de la hoja a limpiar (o presione Enter para usar la primera): ").strip()
    if not sheet_name:
        sheet_name = sheets[0]
    elif sheet_name not in sheets:
        # Try case insensitive match or index match
        matched = False
        for s in sheets:
            if s.lower() == sheet_name.lower():
                sheet_name = s
                matched = True
                break
        if not matched:
            try:
                sheet_idx = int(sheet_name) - 1
                if 0 <= sheet_idx < len(sheets):
                    sheet_name = sheets[sheet_idx]
                    matched = True
            except ValueError:
                pass
        if not matched:
            print(f"Error: La hoja '{sheet_name}' no existe en el archivo.")
            return

    print(f"\nProcesando hoja: {sheet_name}")
    
    # Read sheet
    try:
        df = pd.read_excel(target_file, sheet_name=sheet_name)
    except Exception as e:
        print(f"Error al cargar la hoja: {e}")
        return
        
    required_cols = ["TARJETA", "CUENTA", "DNI", "CLIENTE", "TELEFONO"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        print(f"Error: No se encontraron las siguientes columnas requeridas: {missing}")
        return

    print("Columnas encontradas con éxito.")
    
    alerts = []
    clean_indices = []
    
    # Store cleaned columns as lists to update df
    cleaned_tarjeta = []
    cleaned_cuenta = []
    cleaned_dni_list = []
    cleaned_cliente = []
    cleaned_telefono = []
    
    # Find the assessor column (handling case and spaces)
    asesor_col = next((c for c in df.columns if str(c).strip().upper() == "ASESOR"), None)
    
    anuladas_count = 0
    for idx, row in df.iterrows():
        # Omitir registros si ASESOR es ANULADA
        if asesor_col is not None:
            raw_asesor = row[asesor_col]
            if not pd.isna(raw_asesor) and str(raw_asesor).strip().upper() == "ANULADA":
                anuladas_count += 1
                continue
                
        # Get raw values
        raw_tarjeta = row["TARJETA"]
        raw_cuenta = row["CUENTA"]
        raw_dni = row["DNI"]
        raw_cliente = row["CLIENTE"]
        raw_telefono = row["TELEFONO"]
        
        # Clean values
        c_tarjeta = clean_card(raw_tarjeta)
        c_cuenta = clean_account(raw_cuenta)
        c_dni = clean_dni(raw_dni)
        c_cliente = str(raw_cliente).strip() if not pd.isna(raw_cliente) else ""
        c_telefono = clean_phone(raw_telefono)
        
        # Validation checks
        failed_fields = []
        if c_tarjeta is None:
            failed_fields.append(f"TARJETA ({raw_tarjeta})")
        if c_cuenta is None:
            failed_fields.append(f"CUENTA ({raw_cuenta})")
        if c_dni is None:
            failed_fields.append(f"DNI ({raw_dni})")
        if not c_cliente:
            failed_fields.append(f"CLIENTE ({raw_cliente})")
        if c_telefono is None:
            failed_fields.append(f"TELEFONO ({raw_telefono})")
            
        if failed_fields:
            # Format row representation for the alert
            dni_display = str(raw_dni).strip() if not pd.isna(raw_dni) else "SIN DNI"
            if dni_display.endswith('.0'):
                dni_display = dni_display[:-2]
            alerts.append({
                "fila": idx + 2, # Excel row index
                "dni": dni_display,
                "cliente": c_cliente or str(raw_cliente),
                "detalles": ", ".join(failed_fields)
            })
        else:
            clean_indices.append(idx)
            cleaned_tarjeta.append(c_tarjeta)
            cleaned_cuenta.append(c_cuenta)
            cleaned_dni_list.append(c_dni)
            cleaned_cliente.append(c_cliente)
            cleaned_telefono.append(c_telefono)
            
    # Filter and update dataframe for BASE PROLIJA
    df_clean = df.loc[clean_indices].copy()
    df_clean["TARJETA"] = cleaned_tarjeta
    df_clean["CUENTA"] = cleaned_cuenta
    df_clean["DNI"] = cleaned_dni_list
    df_clean["CLIENTE"] = cleaned_cliente
    df_clean["TELEFONO"] = cleaned_telefono
    
    # Ensure all cleaned columns are treated as strings to avoid scientific notation
    for col in ["TARJETA", "CUENTA", "DNI", "TELEFONO"]:
        df_clean[col] = df_clean[col].astype(str)
        
    # Create the folder DEVUELTO if it doesn't exist
    output_dir = os.path.join(search_dir, "DEVUELTO")
    os.makedirs(output_dir, exist_ok=True)
    
    base_prolija_path = os.path.join(output_dir, "BASE PROLIJA.xlsx")
    predictivo_path = os.path.join(output_dir, "subir a predictivo.xlsx")
    
    # Save BASE PROLIJA.xlsx using openpyxl directly to force text format on fields
    try:
        # Write clean dataframe to BASE PROLIJA
        wb_prolija = openpyxl.Workbook()
        ws_prolija = wb_prolija.active
        ws_prolija.title = "Base Limpia"
        
        # Write headers
        ws_prolija.append(list(df_clean.columns))
        
        # Write rows and set format for numeric strings to '@' (text)
        for _, row in df_clean.iterrows():
            row_vals = []
            for col in df_clean.columns:
                val = row[col]
                row_vals.append(to_str_no_sci(val))
            ws_prolija.append(row_vals)
            
        # Format ALL columns as text to avoid any scientific notation
        for col_idx in range(1, len(df_clean.columns) + 1):
            for row_idx in range(2, len(df_clean) + 2):
                cell = ws_prolija.cell(row=row_idx, column=col_idx)
                cell.number_format = '@'  # Force text format
                    
        wb_prolija.save(base_prolija_path)
        print(f"\nArchivo 'BASE PROLIJA.xlsx' guardado exitosamente en: {base_prolija_path}")
    except Exception as e:
        print(f"Error al guardar BASE PROLIJA: {e}")
        
    # Save subir a predictivo.xlsx
    try:
        df_pred = pd.DataFrame({
            "TELEFONO": cleaned_telefono,
            "CUENTA": cleaned_cuenta,
            "DNI": cleaned_dni_list,
            "TARJETA": cleaned_tarjeta,
            "NOMBRE": cleaned_cliente
        })
        
        wb_pred = openpyxl.Workbook()
        ws_pred = wb_pred.active
        ws_pred.title = "Predictivo"
        
        ws_pred.append(list(df_pred.columns))
        for _, row in df_pred.iterrows():
            row_vals = [to_str_no_sci(val) for val in row]
            ws_pred.append(row_vals)
            
        # Format all columns in predictivo as text
        for col_idx in range(1, len(df_pred.columns) + 1):
            for row_idx in range(2, len(df_pred) + 2):
                cell = ws_pred.cell(row=row_idx, column=col_idx)
                cell.number_format = '@'
                
        wb_pred.save(predictivo_path)
        print(f"Archivo 'subir a predictivo.xlsx' guardado exitosamente en: {predictivo_path}")
    except Exception as e:
        print(f"Error al guardar subir a predictivo: {e}")
        
    if anuladas_count > 0:
        print(f"\nSe omitieron {anuladas_count} registros con ASESOR = 'ANULADA'.")

    # Print alerts/errors
    print(f"\n=== REPORTES DE ALERTAS ({len(alerts)} filas no procesadas) ===")
    if alerts:
        print(f"{'Fila Excel':<12} | {'DNI':<15} | {'Cliente':<25} | {'Campos Fallidos'}")
        print("-" * 80)
        for alert in alerts:
            print(f"{alert['fila']:<12} | {alert['dni']:<15} | {alert['cliente'][:23]:<25} | {alert['detalles']}")
    else:
        print("¡Todas las filas se limpiaron correctamente!")

if __name__ == "__main__":
    main()
