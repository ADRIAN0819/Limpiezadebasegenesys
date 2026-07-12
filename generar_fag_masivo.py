"""
Genera la plantilla FAG MASIVO a partir de los registros de la carpeta LIMPIAR
cuyo ASESOR sea ANULADA o NO EXISTE.
"""

import openpyxl
import os
import re
import pandas as pd
import datetime

# ── Rutas ────────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LIMPIAR_DIR = os.path.join(BASE_DIR, "LIMPIAR")
MODELO_PATH = os.path.join(LIMPIAR_DIR, "FAG SUBIDA MODELO.xlsx")

print("=" * 60)
print("   GENERADOR DE PLANTILLA FAG MASIVO (ANULADAS / NO EXISTE)")
print("=" * 60)
print()

# ── 1. Buscar archivos disponibles en la carpeta LIMPIAR ─────────────────────
if not os.path.exists(LIMPIAR_DIR):
    print(f"[!] La carpeta {LIMPIAR_DIR} no existe.")
    exit(1)

xlsx_files = [f for f in os.listdir(LIMPIAR_DIR) if f.endswith(".xlsx") and f != "FAG SUBIDA MODELO.xlsx"]
if not xlsx_files:
    print("[!] No se encontraron archivos de gestión (.xlsx) en la carpeta LIMPIAR.")
    exit(1)

print("Archivos disponibles en la carpeta LIMPIAR:")
for idx, f in enumerate(xlsx_files, start=1):
    print(f"  [{idx}] {f}")

while True:
    try:
        sel = input(f"Selecciona el archivo a procesar (1-{len(xlsx_files)}): ").strip()
        sel_idx = int(sel) - 1
        if 0 <= sel_idx < len(xlsx_files):
            INPUT_FILE = os.path.join(LIMPIAR_DIR, xlsx_files[sel_idx])
            break
    except ValueError:
        pass
    print("  [!] Selección inválida.")

# ── 2. Cargar archivo y seleccionar hoja ──────────────────────────────────────
print(f"\nCargando {os.path.basename(INPUT_FILE)}...")
try:
    wb_in = openpyxl.load_workbook(INPUT_FILE, read_only=True)
    sheets = wb_in.sheetnames
except Exception as e:
    print(f"[!] Error al abrir el archivo: {e}")
    exit(1)

# Buscar hoja por defecto que contenga "LLENAR DATOS"
default_sheet = next((s for s in sheets if "LLENAR DATOS" in s.upper().replace("/", "")), None)

print("\nHojas disponibles:")
for idx, s in enumerate(sheets, start=1):
    suffix = " (Default)" if s == default_sheet else ""
    print(f"  [{idx}] {s}{suffix}")

while True:
    sel = input(f"Selecciona la hoja (Enter para usar el default '{default_sheet}' si existe): ").strip()
    if not sel and default_sheet:
        SHEET_NAME = default_sheet
        break
    try:
        sel_idx = int(sel) - 1
        if 0 <= sel_idx < len(sheets):
            SHEET_NAME = sheets[sel_idx]
            break
    except ValueError:
        pass
    print("  [!] Selección inválida.")

# ── 3. Pedir datos requeridos ────────────────────────────────────────────────
def pedir_fecha():
    now = datetime.datetime.now()
    default_fecha = now.strftime("%d/%m/%Y")
    while True:
        v = input(f"Ingresa la FECHA (DD/MM/AAAA) [Default: {default_fecha}]: ").strip()
        if not v:
            return default_fecha
        if re.match(r"^\d{2}/\d{2}/\d{4}$", v):
            return v
        print("  [!] Formato incorrecto. Usa DD/MM/AAAA.")

def pedir_hora():
    now = datetime.datetime.now()
    default_hora = now.strftime("%H:%M:%S")
    while True:
        v = input(f"Ingresa la HORA (HH:MM:SS) [Default: {default_hora}]: ").strip()
        if not v:
            return default_hora
        if re.match(r"^\d{2}:\d{2}:\d{2}$", v):
            return v
        print("  [!] Formato incorrecto. Usa HH:MM:SS.")

INPUT_FECHA = pedir_fecha()
INPUT_HORA = pedir_hora()
while True:
    INPUT_NOMBRE = input("Ingresa el nombre del archivo final (sin extensión) [Default: FAG_MASIVO]: ").strip()
    if not INPUT_NOMBRE:
        INPUT_NOMBRE = "FAG_MASIVO"
    if INPUT_NOMBRE.lower().endswith(".xlsx"):
        INPUT_NOMBRE = INPUT_NOMBRE[:-5]
    if INPUT_NOMBRE:
        break

# Asegurar que la carpeta DEVUELTO exista
DEVUELTO_DIR = os.path.join(LIMPIAR_DIR, "DEVUELTO")
os.makedirs(DEVUELTO_DIR, exist_ok=True)
OUTPUT_FILE = os.path.join(DEVUELTO_DIR, INPUT_NOMBRE + ".xlsx")

print()
print(f"  Fecha         : {INPUT_FECHA}")
print(f"  Hora          : {INPUT_HORA}")
print(f"  Archivo Salida: {OUTPUT_FILE}")
print()
confirmar = input("¿Confirmar y generar el Excel? (s/n): ").strip().lower()
if confirmar != "s":
    print("Operación cancelada.")
    exit(0)

# ── 4. Procesar Datos ────────────────────────────────────────────────────────
print("\nLeyendo registros...")
df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME)

# Normalizar cabeceras a mayúsculas y sin espacios
df.columns = [str(c).strip().upper() for c in df.columns]

# Encontrar columnas requeridas o alternativas
def get_col_val(row, possible_names):
    for name in possible_names:
        if name in df.columns:
            val = row[name]
            if pd.notna(val):
                return str(val).strip()
    return None

# Filtrar registros donde ASESOR es ANULADA o NO EXISTE
asesor_col = next((c for c in df.columns if "ASESOR" in c), None)
if not asesor_col:
    print("[!] Error: No se encontró la columna ASESOR en la hoja seleccionada.")
    exit(1)

filtered_rows = []
for idx, row in df.iterrows():
    raw_asesor = row[asesor_col]
    if pd.notna(raw_asesor):
        asesor_str = str(raw_asesor).strip().upper()
        if asesor_str in ["ANULADA", "NO EXISTE"]:
            filtered_rows.append((row, asesor_str))

print(f"Se encontraron {len(filtered_rows)} registros con ASESOR = 'ANULADA' o 'NO EXISTE'.")
if not filtered_rows:
    print("No hay registros para generar. Operación finalizada.")
    exit(0)

# ── 5. Escribir plantilla de salida usando el modelo si existe ───────────────
if os.path.exists(MODELO_PATH):
    print("Usando plantilla modelo FAG SUBIDA MODELO.xlsx...")
    wb_out = openpyxl.load_workbook(MODELO_PATH)
    ws_out = wb_out.active
    # Limpiar datos antiguos pero conservar cabeceras
    if ws_out.max_row > 1:
        ws_out.delete_rows(2, ws_out.max_row)
else:
    print("Plantilla modelo no encontrada. Creando nuevo archivo...")
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "REGISTROS WEBFORM"

output_columns = [
    "EXPEDIENTE", "FECHA", "HORA", "TIPO DE GESTIÓN", "GESTIÓN", "GESTIÓN MOTIVO",
    "FECHA Y HORA DE ALERTA", "FECHA Y HORA DE ATENCION", "Of. Banco de la Nación",
    "Anexo Interno", "Nombre Funcionario BN", "CELULAR/TELEFONO", "DNI",
    "NOMBRE DEL CLIENTE", "CORREO", "CUENTA EMISORA", "CUENTA RECEPTORA",
    "CUENTA RECEPTORA 2", "CUENTA RECEPTORA 3", "NRO. GIRO 1", "NRO. GIRO 2",
    "NRO. GIRO 3", "NRO. GIRO 4", "NRO. GIRO 5", "TARJETA", "N° BLQ",
    "FECHA(BLQ_VIG)", "CANAL", "REGLA MONITOREO", "REGLA O PARAMETRO DE BLOQUEO",
    "SITUACION_BDUC", "CALIFICACION", "OPCION CALIFICACION",
    "FECHA Y HORA DE ENVIO DE CORREO", "IMPORTE DE FRAUDE", "GESTOR", "COMENTARIO",
    "DIA DE EVENTO", "HORA DE EVENTO", "TIEMPO DE DURACION", "G2",
    "FALLECIMIENTO (SI/NO)", "APLICACIÓN", "Columna2", "Columna3",
    "RESULTADO DE LLAMADA", "NIVEL DE RESPUESTA", "MOTIVO DE ATENCION",
    "VALIDACION DE IDENTIDAD", "TIPO DE TRANSACCION", "IMPORTE RECUPERADO",
    "NUMERO DE RECLAMO", "TIPO DE FRAUDE", "VIGILANCIA DE CUENTA",
    "LEVANTAMIENTO VIGILANCIA", "SOLUCION DE CASO", "FECHA Y HORA BLOQUEO",
    "FECHA Y HORA DESBLOQUEO", "COMUNICACION 1", "COMUNICACION 2",
    "COMUNICACION 3", "FECHA MODIFICACION", "MATERIALIZACION DE FRAUDE",
    "SALDO DISPONIBLE", "TIPO DE CUENTA", "CONCLUSION", "FECHA BLOQUEO DE TARJETA"
]

# Escribir cabeceras si creamos un nuevo archivo
if not os.path.exists(MODELO_PATH):
    for col_num, col_name in enumerate(output_columns, start=1):
        ws_out.cell(row=1, column=col_num, value=col_name)

def clean_val(v):
    if v is None or pd.isna(v):
        return "-"
    s = str(v).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s if s else "-"

def format_hora(h):
    if not h or pd.isna(h):
        return "-"
    s = str(h).strip()
    if "|" in s:
        s = s.split("|")[0].strip()
    match = re.match(r"(\d{1,2}:\d{2})", s)
    if match:
        return match.group(1)
    return s[:5] if len(s) >= 5 else s

row_out = 2
for row, asesor_str in filtered_rows:
    # Obtener valores limpiando y formateando
    dni = clean_val(get_col_val(row, ["DNI"]))
    nombre = clean_val(get_col_val(row, ["CLIENTE", "NOMBRE DEL CLIENTE", "BASE FINAL[BASE WF.NOMBRE DEL CLIENTE]"]))
    cuenta = clean_val(get_col_val(row, ["CUENTA", "CUENTA EMISORA", "BASE FINAL[BASE WF.CUENTA]"]))
    tarjeta = clean_val(get_col_val(row, ["TARJETA", "BASE FINAL[TARJETA]"]))
    fechatrx = clean_val(get_col_val(row, ["BASE FINAL[FECHATRX]", "FECHATRX", "DIA DE EVENTO"]))
    horatrx_raw = get_col_val(row, ["BASE FINAL[HORATRX]", "HORATRX", "HORA DE EVENTO"])
    celular = clean_val(get_col_val(row, ["TELEFONO", "CELULAR", "CELULAR/TELEFONO", "BASE FINAL[BASE WF.CELULAR/TELEFONO]"]))

    # Extraer Canal del primer campo si aplica
    canal = "-"
    first_col_name = df.columns[0]
    first_val = row[first_col_name]
    if pd.notna(first_val):
        first_val_str = str(first_val).strip()
        if len(first_val_str) > 26:
            canal = first_val_str[26:]
    if canal == "-":
        canal_val = get_col_val(row, ["BASE FINAL[REGLA]", "REGLA MONITOREO", "CANAL"])
        if canal_val:
            canal = canal_val

    hora_evento = format_hora(horatrx_raw)

    # Determinar comentario y comunicacion
    comentario_val = ""
    if asesor_str == "ANULADA":
        comunicacion_1_val = "FAG- TJ ANULADA"
    else:
        comunicacion_1_val = "FAG-TJ NO EXISTE"

    # Construir fila
    row_data = {
        "EXPEDIENTE": "",
        "FECHA": INPUT_FECHA,
        "HORA": INPUT_HORA,
        "TIPO DE GESTIÓN": "Outbound",
        "GESTIÓN": "Call Out Monitoreo",
        "GESTIÓN MOTIVO": "",
        "FECHA Y HORA DE ALERTA": "",
        "FECHA Y HORA DE ATENCION": "",
        "Of. Banco de la Nación": "",
        "Anexo Interno": "",
        "Nombre Funcionario BN": "",
        "CELULAR/TELEFONO": celular,
        "DNI": dni,
        "NOMBRE DEL CLIENTE": nombre,
        "CORREO": "",
        "CUENTA EMISORA": cuenta,
        "CUENTA RECEPTORA": "",
        "CUENTA RECEPTORA 2": "",
        "CUENTA RECEPTORA 3": "",
        "NRO. GIRO 1": "",
        "NRO. GIRO 2": "",
        "NRO. GIRO 3": "",
        "NRO. GIRO 4": "",
        "NRO. GIRO 5": "",
        "TARJETA": tarjeta,
        "N° BLQ": "",
        "FECHA(BLQ_VIG)": "",
        "CANAL": canal,
        "REGLA MONITOREO": "",
        "REGLA O PARAMETRO DE BLOQUEO": "",
        "SITUACION_BDUC": "ACTUALIZADO",
        "CALIFICACION": "FAG - Fraude por análisis del gestor",
        "OPCION CALIFICACION": "",
        "FECHA Y HORA DE ENVIO DE CORREO": "",
        "IMPORTE DE FRAUDE": "",
        "GESTOR": "DAVID JOSUE ZAMBRANO LEON",
        "COMENTARIO": comentario_val,
        "DIA DE EVENTO": fechatrx,
        "HORA DE EVENTO": hora_evento,
        "TIEMPO DE DURACION": "Activa",
        "G2": "",
        "FALLECIMIENTO (SI/NO)": "NO",
        "APLICACIÓN": "",
        "Columna2": "",
        "Columna3": "",
        "RESULTADO DE LLAMADA": "No Contactado",
        "NIVEL DE RESPUESTA": "Cliente no contesta",
        "MOTIVO DE ATENCION": "",
        "VALIDACION DE IDENTIDAD": "",
        "TIPO DE TRANSACCION": "",
        "IMPORTE RECUPERADO": "",
        "NUMERO DE RECLAMO": "",
        "TIPO DE FRAUDE": "",
        "VIGILANCIA DE CUENTA": "",
        "LEVANTAMIENTO VIGILANCIA": "NO APLICA",
        "SOLUCION DE CASO": "Solucionada",
        "FECHA Y HORA BLOQUEO": "",
        "FECHA Y HORA DESBLOQUEO": "",
        "COMUNICACION 1": comunicacion_1_val,
        "COMUNICACION 2": "",
        "COMUNICACION 3": "",
        "FECHA MODIFICACION": "",
        "MATERIALIZACION DE FRAUDE": "",
        "SALDO DISPONIBLE": "",
        "TIPO DE CUENTA": "",
        "CONCLUSION": "",
        "FECHA BLOQUEO DE TARJETA": ""
    }

    for col_num, col_name in enumerate(output_columns, start=1):
        val = row_data.get(col_name, "")
        cell = ws_out.cell(row=row_out, column=col_num, value=val)
        cell.number_format = "@"

    row_out += 1

wb_out.save(OUTPUT_FILE)
print(f"\n[OK] Plantilla FAG MASIVO generada en: {OUTPUT_FILE}")
print(f"     Total registros omitidos procesados: {row_out - 2}")
