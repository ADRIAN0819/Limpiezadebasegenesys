import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import io
import csv
import numpy as np
from clean_excel import clean_card, clean_account, clean_dni, clean_phone, to_str_no_sci

# Page configuration
st.set_page_config(
    page_title="Herramienta de Gestión de Fraudes - Genesys",
    page_icon="🚨",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom Styling for premium feel
st.markdown("""
    <style>
    .main {
        background-color: #f8f9fa;
    }
    .stButton>button {
        background-color: #4CAF50;
        color: white;
        border-radius: 8px;
        padding: 10px 24px;
        font-weight: bold;
        border: none;
        transition: all 0.3s ease;
    }
    .stButton>button:hover {
        background-color: #45a049;
        transform: scale(1.02);
    }
    .report-card {
        padding: 20px;
        border-radius: 10px;
        background-color: white;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        margin-bottom: 20px;
    }
    h1, h2, h3 {
        color: #2C3E50;
    }
    </style>
""", unsafe_allow_html=True)

# App Header
st.title("🧼 Sistema de Gestión de Bases de Fraude Genesys")
st.markdown("Plataforma interactiva para limpieza de bases y cruce de datos del Predictivo.")

tab1, tab2, tab3, tab4 = st.tabs(["🧼 Limpieza de Base (Excel)", "🔀 Generar Base de Cruce (Predictivo)", "📋 Generar Plantilla (Masivo)", "🔧 Corrección de Cruces"])

# ================= TAB 1: LIMPIEZA DE BASE =================
with tab1:
    st.header("Limpieza y Validación de Excel")
    st.markdown("Valida columnas de tarjetas, cuentas, DNIs, teléfonos y nombres de clientes.")

    # Initialize session state variables if they don't exist
    if "t1_processed" not in st.session_state:
        st.session_state.t1_processed = False
        st.session_state.t1_prolija_data = None
        st.session_state.t1_pred_data = None
        st.session_state.t1_alerts = []
        st.session_state.t1_total_rows = 0
        st.session_state.t1_clean_rows = 0
        st.session_state.t1_anuladas_count = 0
        st.session_state.t1_anuladas_list = []
        st.session_state.t1_current_file = None
        st.session_state.t1_current_sheet = None

    uploaded_file = st.file_uploader("Selecciona el archivo Excel (.xlsx)", type=["xlsx"], key="t1_uploader")

    if uploaded_file is not None:
        try:
            # Load workbook in read-only to list sheet names quickly
            wb_check = openpyxl.load_workbook(uploaded_file, read_only=True)
            sheets = wb_check.sheetnames
            
            # User input / Selection of the sheet
            sheet_name = st.selectbox(
                "Selecciona la hoja a procesar",
                options=sheets,
                index=0,
                key="t1_sheet_select"
            )
            
            # Reset session state if the uploaded file or the sheet name changes
            file_id = uploaded_file.name
            if st.session_state.t1_current_file != file_id or st.session_state.t1_current_sheet != sheet_name:
                st.session_state.t1_processed = False
                st.session_state.t1_prolija_data = None
                st.session_state.t1_pred_data = None
                st.session_state.t1_alerts = []
                st.session_state.t1_total_rows = 0
                st.session_state.t1_clean_rows = 0
                st.session_state.t1_anuladas_count = 0
                st.session_state.t1_anuladas_list = []
                st.session_state.t1_current_file = file_id
                st.session_state.t1_current_sheet = sheet_name
            
            # Process Button
            if st.button("Procesar Archivo", key="t1_process_btn") or st.session_state.t1_processed:
                if not st.session_state.t1_processed:
                    with st.spinner("Procesando y validando base de datos..."):
                        # Reset file pointer and read sheet
                        uploaded_file.seek(0)
                        df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
                        
                        required_cols = ["TARJETA", "CUENTA", "DNI", "CLIENTE", "TELEFONO"]
                        missing = [c for c in required_cols if c not in df.columns]
                        
                        if missing:
                            st.error(f"❌ Error: El archivo no contiene las columnas requeridas: {missing}")
                        else:
                            alerts = []
                            clean_indices = []
                            anuladas_count = 0
                            anuladas_list = []
                            
                            cleaned_tarjeta = []
                            cleaned_cuenta = []
                            cleaned_dni_list = []
                            cleaned_cliente = []
                            cleaned_telefono = []
                            
                            # Find the assessor column (handling case and spaces)
                            asesor_col = next((c for c in df.columns if str(c).strip().upper() == "ASESOR"), None)
                            
                            # Process and validate
                            for idx, row in df.iterrows():
                                # Omitir registros si ASESOR es ANULADA
                                if asesor_col is not None:
                                    raw_asesor = row[asesor_col]
                                    if not pd.isna(raw_asesor) and str(raw_asesor).strip().upper() == "ANULADA":
                                        anuladas_count += 1
                                        anuladas_list.append({
                                            "Fila Excel": idx + 2,
                                            "DNI": str(row.get("DNI", "")).replace(".0", ""),
                                            "Cliente": str(row.get("CLIENTE", "")),
                                            "Tarjeta": str(row.get("TARJETA", "")).replace(".0", ""),
                                            "Cuenta": str(row.get("CUENTA", "")).replace(".0", ""),
                                            "Teléfono": str(row.get("TELEFONO", "")).replace(".0", ""),
                                            "Asesor": str(raw_asesor)
                                        })
                                        continue
                                        
                                raw_tarjeta = row["TARJETA"]
                                raw_cuenta = row["CUENTA"]
                                raw_dni = row["DNI"]
                                raw_cliente = row["CLIENTE"]
                                raw_telefono = row["TELEFONO"]
                                
                                c_tarjeta = clean_card(raw_tarjeta)
                                c_cuenta = clean_account(raw_cuenta)
                                c_dni = clean_dni(raw_dni)
                                c_cliente = str(raw_cliente).strip() if not pd.isna(raw_cliente) else ""
                                c_telefono = clean_phone(raw_telefono)
                                
                                failed_fields = []
                                if c_tarjeta is None:
                                    failed_fields.append(f"TARJETA ({raw_tarjeta})")
                                if c_cuenta is None:
                                    failed_fields.append(f"CUENTA ({raw_cuenta})")
                                if c_dni is None:
                                    failed_fields.append(f"DNI ({raw_dni})")
                                if not c_cliente:
                                    failed_fields.append(f"CLIENTE ({raw_cliente})")
                                if c_telefono is None:
                                    failed_fields.append(f"TELEFONO ({raw_telefono})")
                                    
                                if failed_fields:
                                    dni_display = str(raw_dni).strip() if not pd.isna(raw_dni) else "SIN DNI"
                                    if dni_display.endswith('.0'):
                                        dni_display = dni_display[:-2]
                                    alerts.append({
                                        "Fila Excel": idx + 2,
                                        "DNI": dni_display,
                                        "Cliente": c_cliente or str(raw_cliente),
                                        "Detalles": ", ".join(failed_fields)
                                    })
                                else:
                                    clean_indices.append(idx)
                                    cleaned_tarjeta.append(c_tarjeta)
                                    cleaned_cuenta.append(c_cuenta)
                                    cleaned_dni_list.append(c_dni)
                                    cleaned_cliente.append(c_cliente)
                                    cleaned_telefono.append(c_telefono)
                                    
                            # Filter and update dataframe for BASE PROLIJA
                            df_clean = df.loc[clean_indices].copy()
                            df_clean["TARJETA"] = cleaned_tarjeta
                            df_clean["CUENTA"] = cleaned_cuenta
                            df_clean["DNI"] = cleaned_dni_list
                            df_clean["CLIENTE"] = cleaned_cliente
                            df_clean["TELEFONO"] = cleaned_telefono
                            
                            for col in ["TARJETA", "CUENTA", "DNI", "TELEFONO"]:
                                df_clean[col] = df_clean[col].astype(str)
                                
                            # 1. Generate BASE PROLIJA in memory
                            prolija_buffer = io.BytesIO()
                            wb_prolija = openpyxl.Workbook()
                            ws_prolija = wb_prolija.active
                            ws_prolija.title = "Base Limpia"
                            
                            ws_prolija.append(list(df_clean.columns))
                            for _, row in df_clean.iterrows():
                                row_vals = []
                                for col in df_clean.columns:
                                    val = row[col]
                                    row_vals.append(to_str_no_sci(val))
                                ws_prolija.append(row_vals)
                                
                            for col_idx in range(1, len(df_clean.columns) + 1):
                                for row_idx in range(2, len(df_clean) + 2):
                                    cell = ws_prolija.cell(row=row_idx, column=col_idx)
                                    cell.number_format = '@'
                                    
                            wb_prolija.save(prolija_buffer)
                            prolija_data = prolija_buffer.getvalue()
                            
                            # 2. Generate subir a predictivo in memory
                            df_pred = pd.DataFrame({
                                "TELEFONO": cleaned_telefono,
                                "CUENTA": cleaned_cuenta,
                                "DNI": cleaned_dni_list,
                                "TARJETA": cleaned_tarjeta,
                                "NOMBRE": cleaned_cliente
                            })
                            
                            pred_buffer = io.BytesIO()
                            wb_pred = openpyxl.Workbook()
                            ws_pred = wb_pred.active
                            ws_pred.title = "Predictivo"
                            
                            ws_pred.append(list(df_pred.columns))
                            for _, row in df_pred.iterrows():
                                row_vals = [to_str_no_sci(val) for val in row]
                                ws_pred.append(row_vals)
                                
                            for col_idx in range(1, len(df_pred.columns) + 1):
                                for row_idx in range(2, len(df_pred) + 2):
                                    cell = ws_pred.cell(row=row_idx, column=col_idx)
                                    cell.number_format = '@'
                                    
                            wb_pred.save(pred_buffer)
                            pred_data = pred_buffer.getvalue()
                            
                            # Store in session state
                            st.session_state.t1_prolija_data = prolija_data
                            st.session_state.t1_pred_data = pred_data
                            st.session_state.t1_alerts = alerts
                            st.session_state.t1_total_rows = len(df)
                            st.session_state.t1_clean_rows = len(df_clean)
                            st.session_state.t1_anuladas_count = anuladas_count
                            st.session_state.t1_anuladas_list = anuladas_list
                            st.session_state.t1_processed = True
                
                # Render results
                if st.session_state.t1_processed:
                    st.success("✅ ¡Procesamiento completado con éxito!")
                    
                    # Metrics
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("Total Registros", st.session_state.t1_total_rows)
                    with col2:
                        st.metric("Registros Limpios (Cruzaron)", st.session_state.t1_clean_rows)
                    with col3:
                        st.metric("Registros Fallidos (No Cruzaron)", len(st.session_state.t1_alerts))
                    with col4:
                        st.metric("Registros Anulados (Omitidos)", st.session_state.t1_anuladas_count)
                        
                    # Download Buttons
                    st.subheader("📥 Descargar Archivos Generados")
                    dl_col1, dl_col2 = st.columns(2)
                    with dl_col1:
                        st.download_button(
                            label="Descargar BASE PROLIJA.xlsx",
                            data=st.session_state.t1_prolija_data,
                            file_name="BASE PROLIJA.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="t1_dl_prolija"
                        )
                    with dl_col2:
                        st.download_button(
                            label="Descargar subir a predictivo.xlsx",
                            data=st.session_state.t1_pred_data,
                            file_name="subir a predictivo.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="t1_dl_pred"
                        )
                        
                    # Display Alerts (failed rows)
                    st.subheader("⚠️ Reporte de Registros que NO Cruzaron")
                    if st.session_state.t1_alerts:
                        df_alerts = pd.DataFrame(st.session_state.t1_alerts)
                        st.dataframe(df_alerts, use_container_width=True)
                    else:
                        st.info("¡Excelente! Todas las filas cruzaron y se limpiaron correctamente.")
                        
                    # Display Cancelled Rows (Asesor = ANULADA)
                    st.subheader("🚫 Reporte de Registros ANULADOS (Omitidos)")
                    if st.session_state.t1_anuladas_list:
                        df_anuladas = pd.DataFrame(st.session_state.t1_anuladas_list)
                        st.dataframe(df_anuladas, use_container_width=True)
                    else:
                        st.info("No se encontraron registros anulados por asesor.")
                        
        except Exception as e:
            st.error(f"Error al leer el archivo Excel: {e}")
    else:
        st.info("💡 Por favor, sube un archivo Excel para comenzar.")

# ================= TAB 2: GENERAR BASE DE CRUCE =================
with tab2:
    st.header("Generar Base de Cruce (Predictivo)")
    st.markdown("Realiza el cruce masivo entre listas de alertas, detalles de interacciones Genesys y bases de gestión pendientes.")

    # Initialize session state for Tab 2
    if "t2_processed" not in st.session_state:
        st.session_state.t2_processed = False
        st.session_state.t2_cruce_data = None
        st.session_state.t2_masivo_data = None
        st.session_state.t2_total_final = 0
        st.session_state.t2_total_masivo = 0

    col_files1, col_files2 = st.columns(2)
    with col_files1:
        uploaded_lista = st.file_uploader("Sube la Lista de Alertas (CSV)", type=["csv"], key="t2_lista")
        uploaded_interacc = st.file_uploader("Sube los Detalles de Interacciones (CSV)", type=["csv"], key="t2_interacc")
        
    with col_files2:
        tipo_base_pend = st.radio(
            "Selecciona el tipo de base pendiente:",
            options=["BASE PEND LLENAR (Con cruce de datos)", "BASE PEND LISTA (Directa)"],
            index=0,
            key="t2_tipo_base_pend"
        )
        
        if tipo_base_pend == "BASE PEND LLENAR (Con cruce de datos)":
            uploaded_pend_llenar = st.file_uploader("Sube BASE PEND LLENAR (xlsx)", type=["xlsx"], key="t2_pend_llenar")
            uploaded_datos = st.file_uploader("Sube LLENADO DE DATOS (xlsx)", type=["xlsx"], key="t2_datos")
        else:
            uploaded_pend_lista = st.file_uploader("Sube BASE PEND LISTA (xlsx)", type=["xlsx"], key="t2_pend_lista")

    st.subheader("Configuración de Calificación")
    col_conf1, col_conf2 = st.columns(2)
    with col_conf1:
        tipo_calif = st.selectbox(
            "Selecciona tipo de calificación:",
            options=["PENDIENTES", "CIERRE"],
            key="t2_tipo_calif"
        )
        
        if tipo_calif == "CIERRE":
            uploaded_comercios = st.file_uploader("Sube la lista de Comercios Riesgosos (xlsx - Opcional, por defecto usa 'COMERCIOS RIESGOSOS.xlsx')", type=["xlsx"], key="t2_comercios")
        else:
            uploaded_comercios = None

    with col_conf2:
        valor_comunicacion = st.text_input("Ingresa el valor para COMUNICACION 1:", value="LLAMADA PREDICTIVA", key="t2_comunicacion")
        nombre_archivo_final = st.text_input("Nombre para el archivo final de gestión (sin extensión):", value="MASIVO_PREDICTIVO", key="t2_nombre_final")

    # Validation check to enable process button
    ready_to_process = False
    if uploaded_lista is not None and uploaded_interacc is not None:
        if tipo_base_pend == "BASE PEND LLENAR (Con cruce de datos)":
            if uploaded_pend_llenar is not None and uploaded_datos is not None:
                ready_to_process = True
        else:
            if uploaded_pend_lista is not None:
                ready_to_process = True

    if ready_to_process:
        if st.button("Generar Cruce Predictivo", key="t2_process_btn") or st.session_state.t2_processed:
            if not st.session_state.t2_processed:
                with st.spinner("Procesando cruce de datos Genesys y Predictivo..."):
                    try:
                        # 1. Load lists and interactions
                        df_lista = pd.read_csv(uploaded_lista, dtype=str)
                        
                        # Fix provisorio por si el archivo de lista viene en una sola columna encerrado en comillas dobles
                        if len(df_lista.columns) == 1 and 'TELEFONO' in df_lista.columns[0]:
                            uploaded_lista.seek(0)
                            df_lista = pd.read_csv(uploaded_lista, dtype=str, quoting=csv.QUOTE_NONE, sep=',')
                            df_lista.columns = [c.replace('"', '') for c in df_lista.columns]
                            for col in df_lista.columns:
                                df_lista[col] = df_lista[col].astype(str).str.replace('"', '', regex=False)
                                
                        # Read interactions (sep=';')
                        df_interacc = pd.read_csv(uploaded_interacc, sep=';', dtype=str)
                        
                        # 2. Orquestar Base Pendiente
                        if tipo_base_pend == "BASE PEND LLENAR (Con cruce de datos)":
                            df_pend_temp = pd.read_excel(uploaded_pend_llenar, dtype=str)
                            df_datos = pd.read_excel(uploaded_datos, dtype=str)
                            
                            df_pend_temp['BASE FINAL[TARJETA]'] = df_pend_temp['BASE FINAL[TARJETA]'].astype(str).str.replace(r'\.0$', '', regex=True)
                            df_datos['BASE FINAL[TARJETA]'] = df_datos['BASE FINAL[TARJETA]'].astype(str).str.replace(r'\.0$', '', regex=True)
                            df_datos_unica = df_datos.drop_duplicates(subset=['BASE FINAL[TARJETA]'], keep='first')
                            
                            columnas_extraer = ['BASE FINAL[TARJETA]', 'TELEFONO', 'DNI', 'CUENTA', 'TARJETA', 'CLIENTE']
                            df_datos_unica = df_datos_unica[columnas_extraer]
                            
                            df_resultado = pd.merge(df_pend_temp, df_datos_unica, on='BASE FINAL[TARJETA]', how='left', suffixes=('', '_datos'))
                            
                            df_resultado['BASE FINAL[BASE WF.CELULAR/TELEFONO]'] = df_resultado['TELEFONO'].combine_first(df_resultado['BASE FINAL[BASE WF.CELULAR/TELEFONO]'])
                            df_resultado['BASE FINAL[BASE WF.DNI]'] = df_resultado['DNI'].combine_first(df_resultado['BASE FINAL[BASE WF.DNI]'])
                            df_resultado['BASE FINAL[BASE WF.CUENTA]'] = df_resultado['CUENTA'].combine_first(df_resultado['BASE FINAL[BASE WF.CUENTA]'])
                            df_resultado['BASE FINAL[BASE WF.TARJETA]'] = df_resultado['TARJETA'].combine_first(df_resultado['BASE FINAL[BASE WF.TARJETA]'])
                            df_resultado['BASE FINAL[BASE WF.NOMBRE DEL CLIENTE]'] = df_resultado['CLIENTE'].combine_first(df_resultado['BASE FINAL[BASE WF.NOMBRE DEL CLIENTE]'])
                            
                            for col in ['BASE FINAL[BASE WF.CELULAR/TELEFONO]', 'BASE FINAL[BASE WF.CUENTA]', 'BASE FINAL[BASE WF.DNI]', 'BASE FINAL[BASE WF.TARJETA]']:
                                df_resultado[col] = df_resultado[col].fillna('').astype(str).str.replace(r'\.0$', '', regex=True).replace('nan', '')
                                
                            df_resultado.drop(columns=['TELEFONO', 'DNI', 'CUENTA', 'TARJETA', 'CLIENTE'], inplace=True)
                            df_pend = df_resultado
                        else:
                            df_pend = pd.read_excel(uploaded_pend_lista, dtype=str)

                        # 3. Process llaves
                        df_lista = df_lista.fillna("")
                        df_lista['llave PEND_temp'] = df_lista['TELEFONO'] + df_lista['CUENTA'] + df_lista['DNI'] + df_lista['TARJETA']
                        
                        df_pend_temp_key = df_pend.copy()
                        df_pend_temp_key['llave PEND_temp'] = df_pend_temp_key['BASE FINAL[BASE WF.CELULAR/TELEFONO]'] + df_pend_temp_key['BASE FINAL[BASE WF.CUENTA]'] + df_pend_temp_key['BASE FINAL[BASE WF.DNI]'] + df_pend_temp_key['BASE FINAL[BASE WF.TARJETA]']
                        df_pend_temp_key['hora_limpia_temp'] = df_pend_temp_key['BASE FINAL[HORATRX]'].astype(str).str.extract(r'(\d{2}:\d{2})')[0].fillna("")
                        df_pend_temp_key = df_pend_temp_key.drop_duplicates(subset=['llave PEND_temp', 'hora_limpia_temp'], keep='first')
                        
                        df_lista = pd.merge(df_lista, df_pend_temp_key[['llave PEND_temp', 'BASE FINAL[HORATRX]']], on='llave PEND_temp', how='left')
                        hora_transaccion_lista = df_lista['BASE FINAL[HORATRX]'].astype(str).str.extract(r'(\d{2}:\d{2})')[0].fillna("")
                        df_lista = df_lista.drop(columns=['BASE FINAL[HORATRX]'])
                        
                        df_lista['llave PEND'] = df_lista['llave PEND_temp'] + hora_transaccion_lista
                        
                        call_record_str = df_lista['CallRecordLastAttempt-TELEFONO'].astype(str)
                        call_record_trunc = call_record_str.str[:16]
                        call_record_dates = pd.to_datetime(call_record_trunc, format='%Y-%m-%dT%H:%M', errors='coerce')
                        call_record_dates_lima = call_record_dates - pd.Timedelta(hours=5)
                        df_lista['FECHA Y HORA DE LISTA PRED'] = call_record_dates_lima.dt.strftime('%d/%m/%Y %H:%M:%S').fillna("")
                        
                        df_lista['LLAVE GENESYS'] = df_lista['inin-outbound-id'] + df_lista['FECHA Y HORA DE LISTA PRED']
                        
                        dt_fin_interacc = pd.to_datetime(df_interacc['Fecha de finalización'], format='%d/%m/%y %H:%M', errors='coerce')
                        dt_fin_interacc_str = dt_fin_interacc.dt.strftime('%d/%m/%Y %H:%M:%S').fillna("")
                        df_interacc['LLAVE GENESYS'] = df_interacc['Identificación de contacto'].fillna("") + dt_fin_interacc_str
                        
                        df_interacc_unica = df_interacc.drop_duplicates(subset=['LLAVE GENESYS'], keep='first')
                        
                        df_lista = pd.merge(df_lista, df_interacc_unica[['LLAVE GENESYS', 'Usuarios - Alertados', 'Fecha de finalización']], 
                                            on='LLAVE GENESYS', how='left')
                        
                        df_lista['AG'] = df_lista['Usuarios - Alertados'].fillna("ZLAO")
                        df_lista['AG'] = df_lista['AG'].replace("", "ZLAO")
                        
                        dt_reg_inter = pd.to_datetime(df_lista['Fecha de finalización'], format='%d/%m/%y %H:%M', errors='coerce')
                        fecha_genesys = pd.to_datetime(df_lista['FECHA Y HORA DE LISTA PRED'], format='%d/%m/%Y %H:%M:%S', errors='coerce')
                        dt_reg_final = dt_reg_inter.combine_first(fecha_genesys)
                        
                        df_lista['FECHA  - REGISTRO'] = dt_reg_final.dt.strftime('%d/%m/%Y').fillna("")
                        df_lista['HORA - REGISTRO'] = dt_reg_final.dt.strftime('%H:%M:%S').fillna("")
                        
                        df_pend = df_pend.fillna("")
                        hora_transaccion_limpia = df_pend['BASE FINAL[HORATRX]'].astype(str).str.extract(r'(\d{2}:\d{2})')[0].fillna("")
                        df_pend['llave PEND'] = df_pend['BASE FINAL[BASE WF.CELULAR/TELEFONO]'] + df_pend['BASE FINAL[BASE WF.CUENTA]'] + df_pend['BASE FINAL[BASE WF.DNI]'] + df_pend['BASE FINAL[BASE WF.TARJETA]'] + hora_transaccion_limpia
                        
                        df_pend_unica = df_pend.drop_duplicates(subset=['llave PEND'], keep='first')
                        
                        df_lista = pd.merge(df_lista, df_pend_unica[['llave PEND', 'BASE FINAL[FECHATRX]', 'BASE FINAL[HORATRX]', 'BASE FINAL[REGLA]', 'BASE FINAL[NOMBRECOMERCIO]']], 
                                            on='llave PEND', how='left')
                        
                        df_lista['fecha trx'] = df_lista['BASE FINAL[FECHATRX]'].fillna("")
                        df_lista['hora trx'] = df_lista['BASE FINAL[HORATRX]'].astype(str).str.extract(r'(\d{2}:\d{2})')[0]
                        df_lista['hora trx'] = df_lista['hora trx'].fillna("")
                        
                        mapa_canal = {
                            'FRM': 'FRM - FRM',
                            'IZIPAY': 'MC - MasterCard (TC)',
                            'SQL_BATCH_BLBM_0054': 'BM - Banca Móvil (App)',
                            'SQL_NRT_BM_0052': 'BM - Banca Móvil (App)',
                            'SQL_NRT_HHA_0047': 'INT-HBK - Multired Virtual (HBK)',
                            'SQL_NRT_ATMH_0072': "ATM - ATM's Multired",
                            'VRM': 'VRM - VISA (TMGDV)',
                            'SQL_BATCH_HCS_0057': 'INT-HBK - Multired Virtual (HBK)',
                            'SQL_BATCH_BLMP_0058': 'BATCH - MICROPAGOS - BATCH - MICROPAGOS',
                            'SQL_BATCH_VN7_0061': 'BATCH VISA N7 - BATCH VISA N7',
                            'SQL-TA-ATM-0002': "ATM - ATM's Multired",
                            'SQL_BATCH_GIRO_0069': 'BATCH – GIROS - BATCH – GIROS',
                            'SQL_BATCH_MTCV_0059': 'MULTICANAL - MULTICANAL'
                        }
                        df_lista['canal'] = df_lista['BASE FINAL[REGLA]'].map(mapa_canal).fillna("")
                        df_lista['DNI8'] = df_lista['DNI'].str.zfill(8)
                        
                        # Qualification Calificacion
                        if tipo_calif == "PENDIENTES":
                            df_lista['CALIFICACION'] = "PEND - Alerta pendiente(Temporal)"
                        else:
                            import os
                            if uploaded_comercios is not None:
                                df_comercios = pd.read_excel(uploaded_comercios, dtype=str)
                                comercios_riesgosos = df_comercios['NOMBRE'].fillna("").astype(str).str.strip().str.upper().tolist()
                            elif os.path.exists("COMERCIOS RIESGOSOS.xlsx"):
                                df_comercios = pd.read_excel("COMERCIOS RIESGOSOS.xlsx", dtype=str)
                                comercios_riesgosos = df_comercios['NOMBRE'].fillna("").astype(str).str.strip().str.upper().tolist()
                            else:
                                comercios_riesgosos = []
                                
                            comercio_pend = df_lista['BASE FINAL[NOMBRECOMERCIO]'].fillna("").astype(str).str.strip().str.upper()
                            es_riesgoso = comercio_pend.isin(comercios_riesgosos)
                            
                            df_lista['CALIFICACION'] = np.where(es_riesgoso, 
                                                               "FAG - Fraude por análisis del gestor", 
                                                               "NFCNC - No fraude cliente no contesta(Definitivo)")
                                                               
                        df_lista['1 COMUNICACIÓN'] = "LLAMADA PREDICTIVA"
                        df_lista['origen'] = "MANUAL"
                        
                        mapa_resp = {
                            'ININ-OUTBOUND-BUSY': 'Cliente no contesta',
                            'ININ-OUTBOUND-DISCONNECT': 'Cliente no contesta',
                            'ININ-OUTBOUND-FAILED-TO-REACH-AGENT': 'Telefono Inoperativo - No recibe llamadas',
                            'ININ-OUTBOUND-FAILED-TO-REACH-FLOW': 'Telefono Inoperativo - No recibe llamadas',
                            'ININ-OUTBOUND-FAX': 'Cliente no contesta',
                            'ININ-OUTBOUND-INVALID-PHONE-NUMBER': 'Telefono no existe',
                            'ININ-OUTBOUND-NO-ANSWER': 'Cliente no contesta',
                            'ININ-OUTBOUND-SIT-CALLABLE': 'Cliente no contesta',
                            'ININ-OUTBOUND-STUCK-INTERACTION': 'Cliente no contesta',
                            'Llamada con Mensaje': 'Buzon de voz',
                            'Default Wrap-up Code': 'Cliente no contesta',
                            'ININ-OUTBOUND-AMBIGUOUS': 'Cliente no contesta',
                            'ININ-OUTBOUND-INTERNAL-ERROR-SKIPPED': 'Cliente no contesta',
                            'ININ-OUTBOUND-SIT-UNCALLABLE': 'Cliente no contesta'
                        }
                        df_lista['nivel de respuesta'] = df_lista['CallRecordLastResult-TELEFONO'].map(mapa_resp).fillna("")
                        
                        # Filter final columns
                        columnas_finales = [
                            "llave PEND", "LLAVE GENESYS", "AG", "FECHA  - REGISTRO", "HORA - REGISTRO", 
                            "fecha trx", "hora trx", "canal", "DNI8", "FECHA Y HORA DE LISTA PRED", 
                            "CALIFICACION", "1 COMUNICACIÓN", "nivel de respuesta", "origen", 
                            "inin-outbound-id", "TELEFONO", "CUENTA", "DNI", "TARJETA", "NOMBRE", 
                            "ContactCallable", "ContactableByVoice", "ContactableBySms", "ContactableByEmail", 
                            "CallRecordLastAttempt-TELEFONO", "CallRecordLastResult-TELEFONO", "CallRecordLastAgentWrapup-TELEFONO", 
                            "SmsLastAttempt-TELEFONO", "SmsLastResult-TELEFONO", "Callable-TELEFONO", 
                            "ContactableByVoice-TELEFONO", "ContactableBySms-TELEFONO", "ContactableByWhatsApp", 
                            "CallRecordLastAttemptCampaign-TELEFONO"
                        ]
                        
                        if 'SmsLastAttemptCampaign-TELEFONO' in df_lista.columns:
                            columnas_finales.append('SmsLastAttemptCampaign-TELEFONO')
                            
                        df_final = df_lista[columnas_finales]
                        
                        # Generate masivo dataframe
                        df_masivo = pd.DataFrame()
                        df_masivo['EXPEDIENTE'] = [""] * len(df_final)
                        df_masivo['FECHA'] = df_final['FECHA  - REGISTRO']
                        df_masivo['HORA'] = df_final['HORA - REGISTRO']
                        df_masivo['TIPO DE GESTIÓN'] = "Outbound"
                        df_masivo['GESTIÓN'] = "Call Out Monitoreo"
                        df_masivo['GESTIÓN MOTIVO'] = ""
                        df_masivo['FECHA Y HORA DE ALERTA'] = ""
                        df_masivo['FECHA Y HORA DE ATENCION'] = ""
                        df_masivo['Of. Banco de la Nación'] = ""
                        df_masivo['Anexo Interno'] = ""
                        df_masivo['Nombre Funcionario BN'] = ""
                        df_masivo['CELULAR/TELEFONO'] = df_final['TELEFONO']
                        df_masivo['DNI'] = df_final['DNI']
                        df_masivo['NOMBRE DEL CLIENTE'] = df_final['NOMBRE']
                        df_masivo['CORREO'] = ""
                        df_masivo['CUENTA EMISORA'] = df_final['CUENTA']
                        df_masivo['CUENTA RECEPTORA'] = ""
                        df_masivo['CUENTA RECEPTORA 2'] = ""
                        df_masivo['CUENTA RECEPTORA 3'] = ""
                        df_masivo['NRO. GIRO 1'] = ""
                        df_masivo['NRO. GIRO 2'] = ""
                        df_masivo['NRO. GIRO 3'] = ""
                        df_masivo['NRO. GIRO 4'] = ""
                        df_masivo['NRO. GIRO 5'] = ""
                        df_masivo['TARJETA'] = df_final['TARJETA']
                        df_masivo['N° BLQ'] = ""
                        df_masivo['FECHA(BLQ_VIG)'] = ""
                        df_masivo['CANAL'] = df_final['canal']
                        df_masivo['REGLA MONITOREO'] = ""
                        df_masivo['REGLA O PARAMETRO DE BLOQUEO'] = ""
                        df_masivo['SITUACION_BDUC'] = "ACTUALIZADO"
                        df_masivo['CALIFICACION'] = df_final['CALIFICACION']
                        df_masivo['OPCION CALIFICACION'] = ""
                        df_masivo['FECHA Y HORA DE ENVIO DE CORREO'] = ""
                        df_masivo['IMPORTE DE FRAUDE'] = ""
                        
                        mapa_gestor = {
                            'ZLAO': 'DAVID JOSUE ZAMBRANO LEON',
                            'AG0008 David Zambrano': 'DAVID JOSUE ZAMBRANO LEON',
                            'AG0024 Antonia Montoya': 'ANTONIA MONTOYA HUAMANI',
                            'AG0108 MIGUEL NAVARRO': 'MIGUEL ANGEL NAVARRO SALAZAR',
                            'AG0124 TERESA CABANILLAS': 'TERESA DEL CARMEN CABANILLAS JAVIER',
                            'AG0126 JOSUE OLAYA': 'JOSUE MANUEL OLAYA DOMÍNGUEZ',
                            'AG0129 JANICE MENDOZA': 'JANICE MARJURIE MENDOZA ZAVALA',
                            'AG0130 RAQUEL MARIANO': 'RAQUEL AYME MARIANO ORIHUELA',
                            'AG0134 MANUEL PEÑA': 'MANUEL MARTIN PEÑA BENITES',
                            'AG0142 Sandra Serrano': 'SANDRA SERRANO MOLINA',
                            'AG0151 YOSEP SANGAY': 'YOSEP SANGAY VEGA',
                            'AG0152 VALERIA ESPINOZA': 'VALERIA BELEN ESPINOZA DIAZ',
                            'AG0176 MAURICIO SALCEDO': 'MAURICIO EDSON SALCEDO ENRIQUEZ',
                            'AG0197 Lizbeth Chavez': 'LIZBETH CHAVEZ TORRES',
                            'AG0215 ROSARIO HUATUCO': 'ROSARIO DEL PILAR HUATUCO QUISPE',
                            'AG0263 DERECK MINAYA': 'DERECK ENRIQUE MINAYA CHU'
                        }
                        
                        def mapear_gestor(val):
                            val = str(val).strip()
                            for k, v in mapa_gestor.items():
                                if k in val:
                                    return v
                            return '///'
                            
                        df_masivo['GESTOR'] = df_final['AG'].apply(mapear_gestor)
                        df_masivo['COMENTARIO'] = ""
                        df_masivo['DIA DE EVENTO'] = df_final['fecha trx']
                        df_masivo['HORA DE EVENTO'] = df_final['hora trx']
                        df_masivo['TIEMPO DE DURACION'] = "Activa"
                        df_masivo['G2'] = ""
                        df_masivo['FALLECIMIENTO (SI/NO)'] = "NO"
                        df_masivo['APLICACIÓN'] = ""
                        df_masivo['Columna2'] = ""
                        df_masivo['Columna3'] = ""
                        df_masivo['RESULTADO DE LLAMADA'] = "No Contactado"
                        df_masivo['NIVEL DE RESPUESTA'] = df_final['nivel de respuesta']
                        df_masivo['MOTIVO DE ATENCION'] = ""
                        df_masivo['VALIDACION DE IDENTIDAD'] = ""
                        df_masivo['TIPO DE TRANSACCION'] = ""
                        df_masivo['IMPORTE RECUPERADO'] = ""
                        df_masivo['NUMERO DE RECLAMO'] = ""
                        df_masivo['TIPO DE FRAUDE'] = ""
                        df_masivo['VIGILANCIA DE CUENTA'] = ""
                        df_masivo['LEVANTAMIENTO VIGILANCIA'] = "NO APLICA"
                        df_masivo['SOLUCION DE CASO'] = "Solucionada"
                        df_masivo['FECHA Y HORA BLOQUEO'] = ""
                        df_masivo['FECHA Y HORA DESBLOQUEO'] = ""
                        df_masivo['COMUNICACION 1'] = valor_comunicacion
                        df_masivo['COMUNICACION 2'] = ""
                        df_masivo['COMUNICACION 3'] = ""
                        df_masivo['FECHA MODIFICACION'] = ""
                        df_masivo['MATERIALIZACION DE FRAUDE'] = ""
                        df_masivo['SALDO DISPONIBLE'] = ""
                        df_masivo['TIPO DE CUENTA'] = ""
                        df_masivo['CONCLUSION'] = ""
                        df_masivo['FECHA BLOQUEO DE TARJETA'] = ""
                        
                        # Identify invalid rows
                        mascara_invalidos = (df_final['canal'] == "") | (df_final['nivel de respuesta'] == "") | (df_final['hora trx'] == "") | (df_final['fecha trx'] == "") | (df_masivo['CANAL'] == "")
                        
                        df_final = df_final.replace("#N/D", "")
                        df_masivo = df_masivo.replace("#N/D", "")
                        
                        df_masivo = df_masivo[~mascara_invalidos]
                        
                        # Save both dataframes to Excel in memory
                        cruce_buffer = io.BytesIO()
                        df_final.to_excel(cruce_buffer, index=False)
                        cruce_data = cruce_buffer.getvalue()
                        
                        masivo_buffer = io.BytesIO()
                        df_masivo.to_excel(masivo_buffer, index=False)
                        masivo_data = masivo_buffer.getvalue()
                        
                        # Store results in session state
                        st.session_state.t2_cruce_data = cruce_data
                        st.session_state.t2_masivo_data = masivo_data
                        st.session_state.t2_total_final = len(df_final)
                        st.session_state.t2_total_masivo = len(df_masivo)
                        st.session_state.t2_processed = True
                        
                    except Exception as err:
                        st.error(f"Ocurrió un error en el cruce de datos: {err}")
            
            # Show Tab 2 results if processed
            if st.session_state.t2_processed:
                st.success("🎉 ¡Cruce predictivo realizado con éxito!")
                
                col_m1, col_m2 = st.columns(2)
                with col_m1:
                    st.metric("Filas en BASE CRUCE", st.session_state.t2_total_final)
                with col_m2:
                    st.metric("Filas en MASIVO GESTIÓN (Filtradas)", st.session_state.t2_total_masivo)
                    
                st.subheader("📥 Descargar Resultados del Cruce")
                col_dl1, col_dl2 = st.columns(2)
                with col_dl1:
                    st.download_button(
                        label="Descargar BASE CRUCE PRED_NUEVA.xlsx",
                        data=st.session_state.t2_cruce_data,
                        file_name="BASE CRUCE PRED_NUEVA.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="t2_dl_cruce"
                    )
                with col_dl2:
                    st.download_button(
                        label=f"Descargar {nombre_archivo_final}.xlsx",
                        data=st.session_state.t2_masivo_data,
                        file_name=f"{nombre_archivo_final}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="t2_dl_masivo"
                    )
    else:
        st.info("💡 Sube todos los archivos obligatorios arriba para habilitar el procesamiento del cruce predictivo.")

# ================= TAB 3: GENERAR PLANTILLA (MASIVO) =================
with tab3:
    st.header("Generar Plantilla de Gestión de Fraude (Masivo)")
    st.markdown("Genera la plantilla de gestión a partir de la base masiva y los comercios riesgosos.")

    import datetime
    import re
    import os

    if "t3_processed" not in st.session_state:
        st.session_state.t3_processed = False
        st.session_state.t3_output_data = None
        st.session_state.t3_total_rows = 0

    col_t3_files1, col_t3_files2 = st.columns(2)
    with col_t3_files1:
        uploaded_base_masivo = st.file_uploader("Sube BASE MASIVO (xlsx)", type=["xlsx"], key="t3_base_masivo")
    with col_t3_files2:
        uploaded_comercios_t3 = st.file_uploader("Sube COMERCIOS RIESGOSOS (xlsx - Opcional)", type=["xlsx"], key="t3_comercios")

    now = datetime.datetime.now()
    default_fecha = now.strftime("%d/%m/%Y")
    default_hora = now.strftime("%H:%M:%S")

    st.subheader("Configuración de Plantilla")
    col_t3_conf1, col_t3_conf2 = st.columns(2)
    with col_t3_conf1:
        fecha_val = st.text_input("Ingresa la FECHA (DD/MM/AAAA):", value=default_fecha, key="t3_fecha")
        hora_val = st.text_input("Ingresa la HORA (HH:MM:SS):", value=default_hora, key="t3_hora")
    with col_t3_conf2:
        comunicacion_val = st.text_input("Ingresa el valor de COMUNICACION 1:", value="LLAMADA PREDICTIVA", key="t3_comunicacion")
        nombre_archivo_t3 = st.text_input("Nombre para el archivo final (sin extensión):", value="PLANTILLA_MASIVO", key="t3_nombre_final")

    # Reset state if base masivo file changes
    if uploaded_base_masivo is not None:
        if st.session_state.get("t3_current_file") != uploaded_base_masivo.name:
            st.session_state.t3_processed = False
            st.session_state.t3_current_file = uploaded_base_masivo.name

        if st.button("Generar Plantilla Masiva", key="t3_process_btn") or st.session_state.t3_processed:
            if not st.session_state.t3_processed:
                with st.spinner("Procesando y generando plantilla..."):
                    try:
                        # Load COMERCIOS RIESGOSOS
                        comercios_riesgosos = set()
                        if uploaded_comercios_t3 is not None:
                            wb_com = openpyxl.load_workbook(uploaded_comercios_t3)
                            ws_com = wb_com.active
                            for row in ws_com.iter_rows(min_row=2, values_only=True):
                                nombre = row[0]
                                if nombre:
                                    comercios_riesgosos.add(str(nombre).strip().upper())
                        elif os.path.exists("COMERCIOS RIESGOSOS.xlsx"):
                            wb_com = openpyxl.load_workbook("COMERCIOS RIESGOSOS.xlsx")
                            ws_com = wb_com.active
                            for row in ws_com.iter_rows(min_row=2, values_only=True):
                                nombre = row[0]
                                if nombre:
                                    comercios_riesgosos.add(str(nombre).strip().upper())

                        # Load BASE MASIVO
                        wb_base = openpyxl.load_workbook(uploaded_base_masivo)
                        ws_base = wb_base.active

                        # Map headers to indices
                        headers_base = [cell.value for cell in ws_base[1]]
                        def col_idx(name):
                            try:
                                return headers_base.index(name)
                            except ValueError:
                                return None

                        idx_dni        = col_idx("BASE FINAL[BASE WF.DNI]")
                        idx_nombre     = col_idx("BASE FINAL[BASE WF.NOMBRE DEL CLIENTE]")
                        idx_cuenta     = col_idx("BASE FINAL[BASE WF.CUENTA]")
                        idx_tarjeta    = col_idx("BASE FINAL[BASE WF.TARJETA]")
                        idx_canal      = col_idx("BASE FINAL[BASE WF.CANAL]")
                        idx_fechatrx   = col_idx("BASE FINAL[FECHATRX]")
                        idx_horatrx    = col_idx("BASE FINAL[HORATRX]")
                        idx_comercio   = col_idx("BASE FINAL[NOMBRECOMERCIO]")
                        idx_celular    = col_idx("BASE FINAL[BASE WF.CELULAR/TELEFONO]")

                        # Output columns
                        output_columns = [
                            "EXPEDIENTE", "FECHA", "HORA", "TIPO DE GESTIÓN", "GESTIÓN", "GESTIÓN MOTIVO",
                            "FECHA Y HORA DE ALERTA", "FECHA Y HORA DE ATENCION", "Of. Banco de la Nación",
                            "Anexo Interno", "Nombre Funcionario BN", "CELULAR/TELEFONO", "DNI",
                            "NOMBRE DEL CLIENTE", "CORREO", "CUENTA EMISORA", "CUENTA RECEPTORA",
                            "CUENTA RECEPTORA 2", "CUENTA RECEPTORA 3", "NRO. GIRO 1", "NRO. GIRO 2",
                            "NRO. GIRO 3", "NRO. GIRO 4", "NRO. GIRO 5", "TARJETA", "N° BLQ",
                            "FECHA(BLQ_VIG)", "CANAL", "REGLA MONITOREO", "REGLA O PARAMETRO DE BLOQUEO",
                            "SITUACION_BDUC", "CALIFICACION", "OPCION CALIFICACION",
                            "FECHA Y HORA DE ENVIO DE CORREO", "IMPORTE DE FRAUDE", "GESTOR", "COMENTARIO",
                            "DIA DE EVENTO", "HORA DE EVENTO", "TIEMPO DE DURACION", "G2",
                            "FALLECIMIENTO (SI/NO)", "APLICACIÓN", "Columna2", "Columna3",
                            "RESULTADO DE LLAMADA", "NIVEL DE RESPUESTA", "MOTIVO DE ATENCION",
                            "VALIDACION DE IDENTIDAD", "TIPO DE TRANSACCION", "IMPORTE RECUPERADO",
                            "NUMERO DE RECLAMO", "TIPO DE FRAUDE", "VIGILANCIA DE CUENTA",
                            "LEVANTAMIENTO VIGILANCIA", "SOLUCION DE CASO", "FECHA Y HORA BLOQUEO",
                            "FECHA Y HORA DESBLOQUEO", "COMUNICACION 1", "COMUNICACION 2",
                            "COMUNICACION 3", "FECHA MODIFICACION", "MATERIALIZACION DE FRAUDE",
                            "SALDO DISPONIBLE", "TIPO DE CUENTA", "CONCLUSION", "FECHA BLOQUEO DE TARJETA"
                        ]

                        wb_out = openpyxl.Workbook()
                        ws_out = wb_out.active
                        ws_out.title = "REGISTROS WEBFORM"
                        ws_out.sheet_format.defaultRowHeight = 12.75

                        for col_num, col_name in enumerate(output_columns, start=1):
                            ws_out.cell(row=1, column=col_num, value=col_name)

                        def str_val(v):
                            if v is None:
                                return ""
                            return str(v).strip()

                        def format_hora(h):
                            if not h:
                                return ""
                            s = str(h).strip()
                            if "|" in s:
                                s = s.split("|")[0].strip()
                            match = re.match(r"(\d{1,2}:\d{2})", s)
                            if match:
                                return match.group(1)
                            return s[:5] if len(s) >= 5 else s

                        def calificacion(comercio_val):
                            if comercio_val:
                                c = str(comercio_val).strip().upper()
                                if c in comercios_riesgosos:
                                    return "FAG - Fraude por análisis del gestor"
                            return "DAG - Descarte por análisis del gestor"

                        row_out = 2
                        for row in ws_base.iter_rows(min_row=2, values_only=True):
                            dni          = str_val(row[idx_dni])         if idx_dni is not None else ""
                            nombre       = str_val(row[idx_nombre])      if idx_nombre is not None else ""
                            cuenta       = str_val(row[idx_cuenta])      if idx_cuenta is not None else ""
                            tarjeta      = str_val(row[idx_tarjeta])     if idx_tarjeta is not None else ""
                            canal        = str_val(row[idx_canal])       if idx_canal is not None else ""
                            fechatrx     = str_val(row[idx_fechatrx])   if idx_fechatrx is not None else ""
                            horatrx_raw  = str_val(row[idx_horatrx])    if idx_horatrx is not None else ""
                            comercio_val = row[idx_comercio]             if idx_comercio is not None else None
                            celular      = str_val(row[idx_celular])     if idx_celular is not None else ""

                            hora_evento  = format_hora(horatrx_raw)
                            calific      = calificacion(comercio_val)

                            row_data = {
                                "EXPEDIENTE":                    "",
                                "FECHA":                         fecha_val,
                                "HORA":                          hora_val,
                                "TIPO DE GESTIÓN":               "Outbound",
                                "GESTIÓN":                       "Call Out Monitoreo",
                                "GESTIÓN MOTIVO":                "",
                                "FECHA Y HORA DE ALERTA":        "",
                                "FECHA Y HORA DE ATENCION":      "",
                                "Of. Banco de la Nación":        "",
                                "Anexo Interno":                 "",
                                "Nombre Funcionario BN":         "",
                                "CELULAR/TELEFONO":              celular,
                                "DNI":                           dni,
                                "NOMBRE DEL CLIENTE":            nombre,
                                "CORREO":                        "",
                                "CUENTA EMISORA":                cuenta,
                                "CUENTA RECEPTORA":              "",
                                "CUENTA RECEPTORA 2":            "",
                                "CUENTA RECEPTORA 3":            "",
                                "NRO. GIRO 1":                   "",
                                "NRO. GIRO 2":                   "",
                                "NRO. GIRO 3":                   "",
                                "NRO. GIRO 4":                   "",
                                "NRO. GIRO 5":                   "",
                                "TARJETA":                       tarjeta,
                                "N° BLQ":                        "",
                                "FECHA(BLQ_VIG)":                "",
                                "CANAL":                         canal,
                                "REGLA MONITOREO":               "",
                                "REGLA O PARAMETRO DE BLOQUEO":  "",
                                "SITUACION_BDUC":                "ACTUALIZADO",
                                "CALIFICACION":                  calific,
                                "OPCION CALIFICACION":           "",
                                "FECHA Y HORA DE ENVIO DE CORREO": "",
                                "IMPORTE DE FRAUDE":             "",
                                "GESTOR":                        "DAVID JOSUE ZAMBRANO LEON",
                                "COMENTARIO":                    "",
                                "DIA DE EVENTO":                 fechatrx,
                                "HORA DE EVENTO":                hora_evento,
                                "TIEMPO DE DURACION":            "Activa",
                                "G2":                            "",
                                "FALLECIMIENTO (SI/NO)":         "NO",
                                "APLICACIÓN":                    "",
                                "Columna2":                      "",
                                "Columna3":                      "",
                                "RESULTADO DE LLAMADA":          "No Contactado",
                                "NIVEL DE RESPUESTA":            "Cliente no contesta",
                                "MOTIVO DE ATENCION":            "",
                                "VALIDACION DE IDENTIDAD":       "",
                                "TIPO DE TRANSACCION":           "",
                                "IMPORTE RECUPERADO":            "",
                                "NUMERO DE RECLAMO":             "",
                                "TIPO DE FRAUDE":                "",
                                "VIGILANCIA DE CUENTA":          "",
                                "LEVANTAMIENTO VIGILANCIA":      "NO APLICA",
                                "SOLUCION DE CASO":              "Solucionada",
                                "FECHA Y HORA BLOQUEO":          "",
                                "FECHA Y HORA DESBLOQUEO":       "",
                                "COMUNICACION 1":                comunicacion_val,
                                "COMUNICACION 2":                "",
                                "COMUNICACION 3":                "",
                                "FECHA MODIFICACION":            "",
                                "MATERIALIZACION DE FRAUDE":     "",
                                "SALDO DISPONIBLE":              "",
                                "TIPO DE CUENTA":                "",
                                "CONCLUSION":                    "",
                                "FECHA BLOQUEO DE TARJETA":      "",
                            }

                            for col_num, col_name in enumerate(output_columns, start=1):
                                val  = row_data.get(col_name, "")
                                cell = ws_out.cell(row=row_out, column=col_num, value=val)
                                cell.number_format = "@"

                            row_out += 1

                        out_buffer = io.BytesIO()
                        wb_out.save(out_buffer)
                        st.session_state.t3_output_data = out_buffer.getvalue()
                        st.session_state.t3_total_rows = row_out - 2
                        st.session_state.t3_processed = True
                    except Exception as err:
                        st.error(f"Ocurrió un error al generar la plantilla: {err}")

            if st.session_state.t3_processed:
                st.success(f"🎉 ¡Plantilla generada con éxito! ({st.session_state.t3_total_rows} registros)")
                st.download_button(
                    label=f"Descargar {nombre_archivo_t3}.xlsx",
                    data=st.session_state.t3_output_data,
                    file_name=f"{nombre_archivo_t3}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="t3_dl_button"
                )
    else:
        st.info("💡 Por favor, sube el archivo BASE MASIVO (xlsx) para comenzar.")

# ================= HELPER FUNCTIONS FOR TAB 4 =================
def parse_date_from_text(text):
    if not isinstance(text, str):
        return None
    m1 = re.search(r'\b(\d{2})/(\d{2})/(\d{4})\b', text)
    if m1:
        d, m, y = m1.groups()
        return f"{y}-{m}-{d}"
    m2 = re.search(r'\b(\d{4})-(\d{2})-(\d{2})\b', text)
    if m2:
        return m2.group(0)
    return None

def parse_time_from_text(text):
    if not isinstance(text, str):
        return None
    m = re.search(r'\b(\d{1,2}):(\d{2}):(\d{2})\s*(AM|PM|am|pm)?\b', text, re.IGNORECASE)
    if m:
        h = int(m.group(1))
        m_m = m.group(2)
        meridiem = m.group(4)
        if meridiem:
            meridiem = meridiem.upper()
            if meridiem == "PM" and h < 12:
                h += 12
            elif meridiem == "AM" and h == 12:
                h = 0
        h_str = str(h).zfill(2)
        return f"{h_str}:{m_m}"
    m_hm = re.search(r'\b(\d{1,2}):(\d{2})\s*(AM|PM|am|pm)?\b', text, re.IGNORECASE)
    if m_hm:
        h = int(m_hm.group(1))
        m_m = m.group(2)
        meridiem = m.group(3)
        if meridiem:
            meridiem = meridiem.upper()
            if meridiem == "PM" and h < 12:
                h += 12
            elif meridiem == "AM" and h == 12:
                h = 0
        h_str = str(h).zfill(2)
        return f"{h_str}:{m_m}"
    return None

def excel_date_value(date_val):
    if pd.isna(date_val) or date_val == "" or str(date_val).lower() == "nan":
        return ""
    try:
        dt = pd.to_datetime(date_val)
        delta = dt - pd.Timestamp('1899-12-30')
        return str(int(delta.days))
    except Exception as e:
        return ""

def calculate_key_L(row):
    canal = str(row.get('CANAL', '')).strip()
    tarjeta = str(row.get('TARJETA', '')).strip().replace('.0', '')
    dia_de_evento = row.get('DIA DE EVENTO')
    hora_de_evento = str(row.get('HORA DE EVENTO', '')).strip()
    if len(hora_de_evento) > 5:
        hora_de_evento = hora_de_evento[:5]
    regla_visa = str(row.get('REGLA VISA VRM', '')).strip()
    dni = str(row.get('DNI', '')).strip().replace('.0', '')
    cuenta = str(row.get('CUENTA EMISORA', '')).strip().replace('.0', '')
    nombre_cliente = str(row.get('NOMBRE DEL CLIENTE', '')).strip()
    val_an = excel_date_value(dia_de_evento)
    if canal == "BATCH COMERCIOS DIGITALES - BATCH COMERCIOS DIGITALES":
        return f"{tarjeta}{val_an}{canal}"
    elif regla_visa == "SQL_BATCH_HCE_0048" and canal == "Batch - Visa":
        tarjeta_pad = tarjeta if len(tarjeta) == 16 else f"0{tarjeta}"
        return f"{tarjeta_pad}{val_an}{canal}{regla_visa}"
    elif canal == "BM - Banca Móvil (App)":
        return f"{tarjeta}{val_an}{hora_de_evento}"
    elif canal in ["Prestamos Call - Desembolso - Prestamos Call - Desembolso", "Prestamos Call - Autenticacion IVR - Prestamos Call - Autenticacion IVR"]:
        tarjeta_pad = tarjeta.zfill(8) if tarjeta.isdigit() else tarjeta
        return f"{tarjeta_pad}{val_an}{canal}"
    elif canal == "INT-HBK - Multired Virtual (HBK)":
        return f"{tarjeta}{val_an}{hora_de_evento}{canal}"
    elif canal == "VRM - VISA (TMGDV)":
        return f"{tarjeta}{val_an}{canal}"
    elif canal == "MC - MasterCard (TC)":
        return f"{dni}{val_an}MC - MasterCard (TC)"
    elif canal == "FRM - FRM":
        return f"{tarjeta}{canal}"
    elif canal == "Batch VISA N7 - Batch VISA N7":
        tarjeta_mask = f"{tarjeta[:6]}******{tarjeta[-4:]}" if len(tarjeta) >= 10 else tarjeta
        return f"{tarjeta_mask}{val_an}{hora_de_evento}{canal}"
    elif canal == "BATCH – GIROS - BATCH – GIROS":
        try:
            formatted_date = pd.to_datetime(dia_de_evento).strftime('%Y-%m-%d')
        except:
            formatted_date = str(dia_de_evento)
        return f"{nombre_cliente}{formatted_date}{hora_de_evento}{canal}"
    else:
        cuenta_pad = cuenta if len(cuenta) == 11 else f"0{cuenta}"
        return f"{cuenta_pad}{val_an}{canal}"

# ================= TAB 4: CORRECCIÓN DE CRUCES =================
with tab4:
    st.header("🔧 Corrección de Cruces Webform")
    st.markdown("Corrige automáticamente registros de tarjetas, horas y fechas comparándolos con la base de Webform y genera reportes de errores con celdas resaltadas.")

    if "t4_processed" not in st.session_state:
        st.session_state.t4_processed = False
        st.session_state.t4_corregido_data = None
        st.session_state.t4_erroneo_data = None
        st.session_state.t4_total_corregidos = 0
        st.session_state.t4_total_erroneos = 0
        st.session_state.t4_logs = []

    col_t4_1, col_t4_2 = st.columns(2)
    with col_t4_1:
        uploaded_nocruza = st.file_uploader("Sube NOCRUZA (xlsx)", type=["xlsx"], key="t4_nocruza")
    with col_t4_2:
        uploaded_basewf = st.file_uploader("Sube BASE WF (xlsx)", type=["xlsx"], key="t4_basewf")

    if uploaded_nocruza is not None and uploaded_basewf is not None:
        if st.button("Corregir Registros", key="t4_process_btn") or st.session_state.t4_processed:
            if not st.session_state.t4_processed:
                with st.spinner("Procesando y cruzando bases para corrección..."):
                    try:
                        logs = []
                        # Load data
                        df_no = pd.read_excel(uploaded_nocruza)
                        
                        # Rename columns to standard ones to support both NOCRUZA (Banca Móvil) and NOCRUZAHBK formats
                        rename_map = {
                            'REGLA': 'Regla',
                            'FECHA_TRANSACCION': 'FechaTransaccion',
                            'HORA_TRANSACCION': 'HoraTransaccion',
                            'NRO_TARJETA': 'NroTarjeta',
                            'CTA_ORIGEN': 'NroCuenta'
                        }
                        df_no = df_no.rename(columns=rename_map)
                        
                        df_base = pd.read_excel(uploaded_basewf, sheet_name="REGISTROS WEBFORM")
                        
                        # Standarize accounts for matching (lstrip '0' to match both '04754518397' and '4754518397')
                        df_base['CUENTA_str'] = df_base['CUENTA EMISORA'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True).str.lstrip('0')
                        df_no['CUENTA_str'] = df_no['NroCuenta'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True).str.lstrip('0')
                        
                        erroneous_records = []
                        corrected_rows = []
                        
                        for idx, row in df_no.iterrows():
                            acc = row['CUENTA_str']
                            candidates = df_base[df_base['CUENTA_str'] == acc]
                            
                            if len(candidates) == 0:
                                logs.append(f"Fila {idx:<3} | Cuenta {acc:<10} | ERROR: No se encontró en BASE WF")
                                continue
                                
                            no_date = pd.to_datetime(row['FechaTransaccion']).strftime('%Y-%m-%d')
                            no_time = str(row['HoraTransaccion']).strip()
                            if len(no_time) == 5:
                                no_time += ":00"
                            elif len(no_time) > 8:
                                no_time = no_time[:8]
                                
                            no_card = str(row['NroTarjeta']).strip()
                            
                            best_candidate_idx = None
                            best_score = -1
                            
                            for midx, mrow in candidates.iterrows():
                                com = str(mrow['COMUNICACION 1'])
                                score = 0
                                
                                if no_time in com:
                                    score += 10
                                elif no_time[:5] in com:
                                    score += 5
                                if no_date in com:
                                    score += 5
                                formatted_date_slash = pd.to_datetime(no_date).strftime('%d/%m/%Y')
                                if formatted_date_slash in com:
                                    score += 5
                                base_time = str(mrow['HORA DE EVENTO']).strip()
                                if base_time == no_time[:5]:
                                    score += 3
                                base_date = str(mrow['DIA DE EVENTO']).strip()
                                if base_date.startswith(no_date):
                                    score += 2
                                    
                                if score > best_score:
                                    best_score = score
                                    best_candidate_idx = midx
                                    
                            if best_candidate_idx is not None:
                                orig_row = df_base.loc[best_candidate_idx].copy()
                                corrected_row = orig_row.copy()
                                
                                com = str(orig_row['COMUNICACION 1'])
                                
                                regla_no = str(row.get('Regla', '')).upper()
                                expected_canal = orig_row['CANAL']
                                if 'BM' in regla_no:
                                    expected_canal = "BM - Banca Móvil (App)"
                                elif 'HBK' in regla_no:
                                    expected_canal = "INT-HBK - Multired Virtual (HBK)"
                                elif 'VRM' in regla_no:
                                    expected_canal = "VRM - VISA (TMGDV)"
                                    
                                orig_canal = str(orig_row['CANAL']).strip()
                                canal_changed = False
                                if orig_canal != expected_canal:
                                    corrected_row['CANAL'] = expected_canal
                                    canal_changed = True
                                    logs.append(f"F{idx:<3} | {acc:<10} | CANAL           | {orig_canal:<20} | {expected_canal:<20}")
                                
                                orig_card = str(orig_row['TARJETA']).strip().replace('.0', '')
                                last_4_no = no_card[-4:]
                                last_4_orig = orig_card[-4:] if len(orig_card) >= 4 else ""
                                
                                card_changed = False
                                if last_4_no != last_4_orig or orig_card == "0000000000000000" or orig_card == "":
                                    corrected_row['TARJETA'] = no_card
                                    card_changed = True
                                    logs.append(f"F{idx:<3} | {acc:<10} | TARJETA         | {orig_card:<20} | {no_card:<20}")
                                
                                orig_time = str(orig_row['HORA DE EVENTO']).strip()
                                if len(orig_time) > 5:
                                    orig_time = orig_time[:5]
                                    
                                parsed_time = parse_time_from_text(com)
                                correct_time = parsed_time if parsed_time else no_time[:5]
                                if len(correct_time) > 5:
                                    correct_time = correct_time[:5]
                                    
                                time_changed = False
                                if orig_time != correct_time or orig_row['HORA DE EVENTO'] == "00:00" or pd.isna(orig_row['HORA DE EVENTO']):
                                    corrected_row['HORA DE EVENTO'] = correct_time
                                    time_changed = True
                                    logs.append(f"F{idx:<3} | {acc:<10} | HORA DE EVENTO  | {orig_time:<20} | {correct_time:<20}")
                                else:
                                    corrected_row['HORA DE EVENTO'] = correct_time
                                    
                                orig_date_raw = str(orig_row['DIA DE EVENTO']).strip()
                                try:
                                    orig_date = pd.to_datetime(orig_date_raw).strftime('%Y-%m-%d')
                                except:
                                    orig_date = orig_date_raw
                                    
                                parsed_date = parse_date_from_text(com)
                                correct_date = parsed_date if parsed_date else no_date
                                
                                date_changed = False
                                if orig_date != correct_date:
                                    corrected_row['DIA DE EVENTO'] = correct_date
                                    date_changed = True
                                    logs.append(f"F{idx:<3} | {acc:<10} | DIA DE EVENTO   | {orig_date:<20} | {correct_date:<20}")
                                    
                                orig_key = str(orig_row['L']).strip()
                                new_key = calculate_key_L(corrected_row)
                                corrected_row['L'] = new_key
                                if orig_key != new_key:
                                    logs.append(f"F{idx:<3} | {acc:<10} | Llave 'L'       | {orig_key:<20} | {new_key:<20}")
                                    
                                any_changed = card_changed or time_changed or date_changed or canal_changed
                                
                                if any_changed:
                                    obs_list = []
                                    comentarios = []
                                    
                                    if date_changed and time_changed:
                                        obs_list.append("DIA Y HORA DE EVENTO INCORRECTO")
                                        comentarios.append(f"Fecha corregida de '{orig_date}' a '{correct_date}' y Hora de '{orig_time}' a '{correct_time}'")
                                    elif date_changed:
                                        obs_list.append("DIA EVENTO INCORRECTO")
                                        comentarios.append(f"Fecha corregida de '{orig_date}' a '{correct_date}'")
                                    elif time_changed:
                                        obs_list.append("HORA DE EVENTO INCORRECTA")
                                        comentarios.append(f"Hora corregida de '{orig_time}' a '{correct_time}'")
                                        
                                    if card_changed:
                                        obs_list.append("TARJETA INCORRECTA")
                                        comentarios.append(f"Tarjeta corregida de '{orig_card}' a '{no_card}'")
                                        
                                    if canal_changed:
                                        obs_list.append("CANAL INCORRECTO")
                                        comentarios.append(f"Canal corregido de '{orig_canal}' a '{expected_canal}'")
                                    
                                    obs_str = " / ".join(obs_list)
                                    comentario_str = " | ".join(comentarios)
                                    
                                    erroneous_records.append({
                                        'orig_row': orig_row,
                                        'errors': {
                                            'TARJETA': card_changed,
                                            'HORA DE EVENTO': time_changed,
                                            'DIA DE EVENTO': date_changed,
                                            'CANAL': canal_changed
                                        },
                                        'resumen': {
                                            'FECHA DE OPERACION': correct_date,
                                            'CANAL': expected_canal,
                                            'TJ / CTA / DNI': str(orig_row.get('DNI', '')).replace('.0', ''),
                                            'ASESOR ( SEGUN WF)': orig_row.get('GESTOR', ''),
                                            'SUPERVISOR': '',
                                            'OBS': obs_str,
                                            'COMENTARIO ANALISTA': comentario_str
                                        }
                                    })
                                else:
                                    logs.append(f"F{idx:<3} | {acc:<10} | (Correcto)      | Sin cambios en tarjeta/fecha/hora/canal")
                                    
                                if 'CUENTA_str' in corrected_row:
                                    corrected_row = corrected_row.drop('CUENTA_str')
                                corrected_rows.append(corrected_row)
                                
                        # Save Erroneous Records and Resumen
                        erroneo_buffer = io.BytesIO()
                        if len(erroneous_records) > 0:
                            rows_err = []
                            rows_res = []
                            for r in erroneous_records:
                                r_data = r['orig_row'].copy()
                                if 'CUENTA_str' in r_data:
                                    r_data = r_data.drop('CUENTA_str')
                                rows_err.append(r_data)
                                rows_res.append(r['resumen'])
                                
                            df_err = pd.DataFrame(rows_err)
                            df_res = pd.DataFrame(rows_res)
                            
                            with pd.ExcelWriter(erroneo_buffer, engine='openpyxl') as writer:
                                df_err.to_excel(writer, sheet_name="Registros Erroneos", index=False)
                                df_res.to_excel(writer, sheet_name="Resumen Errores", index=False)
                                
                                workbook = writer.book
                                worksheet = workbook["Registros Erroneos"]
                                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                col_map = {col_name: col_idx + 1 for col_idx, col_name in enumerate(df_err.columns)}
                                
                                for row_idx, r in enumerate(erroneous_records):
                                    excel_row = row_idx + 2
                                    for col_name, is_error in r['errors'].items():
                                        if is_error and col_name in col_map:
                                            col_excel_idx = col_map[col_name]
                                            cell = worksheet.cell(row=excel_row, column=col_excel_idx)
                                            cell.fill = red_fill
                        else:
                            with pd.ExcelWriter(erroneo_buffer, engine='openpyxl') as writer:
                                pd.DataFrame(columns=[c for c in df_base.columns if c != 'CUENTA_str']).to_excel(writer, sheet_name="Registros Erroneos", index=False)
                                pd.DataFrame(columns=['FECHA DE OPERACION', 'CANAL', 'TJ / CTA / DNI', 'ASESOR ( SEGUN WF)', 'SUPERVISOR', 'OBS', 'COMENTARIO ANALISTA']).to_excel(writer, sheet_name="Resumen Errores", index=False)
                                
                        erroneo_data = erroneo_buffer.getvalue()
                        
                        # Save Corrected Records
                        df_corrected = pd.DataFrame(corrected_rows)
                        if len(df_corrected) > 0 and 'CUENTA EMISORA' in df_corrected.columns:
                            df_corrected['CUENTA EMISORA'] = df_corrected['CUENTA EMISORA'].apply(
                                lambda x: f"0{str(x).strip().split('.')[0]}" if pd.notna(x) and str(x).strip().split('.')[0].isdigit() and len(str(x).strip().split('.')[0]) == 10 else x
                            )
                        df_corrected_unique = df_corrected.drop_duplicates()
                        
                        df_outbound = df_corrected_unique[df_corrected_unique['TIPO DE GESTIÓN'].astype(str).str.upper().str.contains('OUT')]
                        df_inbound = df_corrected_unique[df_corrected_unique['TIPO DE GESTIÓN'].astype(str).str.upper().str.contains('IN')]
                        
                        corregido_buffer = io.BytesIO()
                        with pd.ExcelWriter(corregido_buffer, engine='openpyxl') as writer:
                            df_outbound.to_excel(writer, sheet_name="OUTBOUND", index=False)
                            df_inbound.to_excel(writer, sheet_name="INBOUND", index=False)
                            
                        corregido_data = corregido_buffer.getvalue()
                        
                        # Store in state
                        st.session_state.t4_corregido_data = corregido_data
                        st.session_state.t4_erroneo_data = erroneo_data
                        st.session_state.t4_total_corregidos = len(df_corrected_unique)
                        st.session_state.t4_total_erroneos = len(erroneous_records)
                        st.session_state.t4_logs = logs
                        st.session_state.t4_processed = True
                        
                    except Exception as err:
                        st.error(f"Error al procesar la corrección: {err}")
                        
            if st.session_state.t4_processed:
                st.success("🎉 ¡Corrección de cruces completada con éxito!")
                
                col_t4_m1, col_t4_m2 = st.columns(2)
                with col_t4_m1:
                    st.metric("Total Registros Corregidos", st.session_state.t4_total_corregidos)
                with col_t4_m2:
                    st.metric("Total Registros Erróneos Detectados", st.session_state.t4_total_erroneos)
                    
                st.subheader("📥 Descargar Archivos de Cruce Corregidos")
                dl_t4_col1, dl_t4_col2 = st.columns(2)
                with dl_t4_col1:
                    st.download_button(
                        label="Descargar devuelto corregido.xlsx",
                        data=st.session_state.t4_corregido_data,
                        file_name="devuelto corregido.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="t4_dl_corregido"
                    )
                with dl_t4_col2:
                    st.download_button(
                        label="Descargar registros erroneos.xlsx",
                        data=st.session_state.t4_erroneo_data,
                        file_name="registros erroneos.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="t4_dl_erroneo"
                    )
                    
                st.subheader("📋 Registro de Procesamiento (Logs)")
                st.code("\n".join(st.session_state.t4_logs), language="text")
    else:
        st.info("💡 Sube ambos archivos obligatorios arriba para habilitar la corrección de cruces.")
