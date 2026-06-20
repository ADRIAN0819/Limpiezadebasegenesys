import pandas as pd
import re
import os
import openpyxl
from openpyxl.styles import PatternFill

def parse_date_from_text(text):
    """
    Tries to find and format a date (DD/MM/YYYY or YYYY-MM-DD) from the text.
    Returns in YYYY-MM-DD format if found, otherwise None.
    """
    if not isinstance(text, str):
        return None
    # Look for DD/MM/YYYY
    m1 = re.search(r'\b(\d{2})/(\d{2})/(\d{4})\b', text)
    if m1:
        d, m, y = m1.groups()
        return f"{y}-{m}-{d}"
    # Look for YYYY-MM-DD
    m2 = re.search(r'\b(\d{4})-(\d{2})-(\d{2})\b', text)
    if m2:
        return m2.group(0)
    return None

def parse_time_from_text(text):
    """
    Tries to find a time (HH:MM:SS or HH:MM) followed optionally by AM/PM from the text.
    Returns HH:MM in 24-hour format if found, otherwise None.
    """
    if not isinstance(text, str):
        return None
    
    # Try HH:MM:SS first
    m = re.search(r'\b(\d{1,2}):(\d{2}):(\d{2})\s*(AM|PM|am|pm)?\b', text, re.IGNORECASE)
    if m:
        h = int(m.group(1))
        m_m = m.group(2)
        meridiem = m.group(4)
        if meridiem:
            meridiem = meridiem.upper()
            if meridiem == "PM" and h < 12:
                h += 12
            elif meridiem == "AM" and h == 12:
                h = 0
        h_str = str(h).zfill(2)
        return f"{h_str}:{m_m}"
        
    # Try HH:MM
    m_hm = re.search(r'\b(\d{1,2}):(\d{2})\s*(AM|PM|am|pm)?\b', text, re.IGNORECASE)
    if m_hm:
        h = int(m_hm.group(1))
        m_m = m.group(2)
        meridiem = m.group(3)
        if meridiem:
            meridiem = meridiem.upper()
            if meridiem == "PM" and h < 12:
                h += 12
            elif meridiem == "AM" and h == 12:
                h = 0
        h_str = str(h).zfill(2)
        return f"{h_str}:{m_m}"
        
    return None

def excel_date_value(date_val):
    """
    Converts YYYY-MM-DD or datetime to Excel date serial number.
    Excel serial date for 2026-06-12 is 46185.
    """
    if pd.isna(date_val) or date_val == "" or str(date_val).lower() == "nan":
        return ""
    try:
        dt = pd.to_datetime(date_val)
        delta = dt - pd.Timestamp('1899-12-30')
        return str(int(delta.days))
    except Exception as e:
        return ""

def calculate_key_L(row):
    """
    Calculates the key 'L' based on the Excel formula.
    """
    canal = str(row.get('CANAL', '')).strip()
    tarjeta = str(row.get('TARJETA', '')).strip().replace('.0', '')
    dia_de_evento = row.get('DIA DE EVENTO')
    # Use only HH:MM for time
    hora_de_evento = str(row.get('HORA DE EVENTO', '')).strip()
    if len(hora_de_evento) > 5:
        hora_de_evento = hora_de_evento[:5]
        
    regla_visa = str(row.get('REGLA VISA VRM', '')).strip()
    dni = str(row.get('DNI', '')).strip().replace('.0', '')
    cuenta = str(row.get('CUENTA EMISORA', '')).strip().replace('.0', '')
    nombre_cliente = str(row.get('NOMBRE DEL CLIENTE', '')).strip()
    
    val_an = excel_date_value(dia_de_evento)
    
    if canal == "BATCH COMERCIOS DIGITALES - BATCH COMERCIOS DIGITALES":
        return f"{tarjeta}{val_an}{canal}"
        
    elif regla_visa == "SQL_BATCH_HCE_0048" and canal == "Batch - Visa":
        tarjeta_pad = tarjeta if len(tarjeta) == 16 else f"0{tarjeta}"
        return f"{tarjeta_pad}{val_an}{canal}{regla_visa}"
        
    elif canal == "BM - Banca Móvil (App)":
        return f"{tarjeta}{val_an}{hora_de_evento}"
        
    elif canal in ["Prestamos Call - Desembolso - Prestamos Call - Desembolso", "Prestamos Call - Autenticacion IVR - Prestamos Call - Autenticacion IVR"]:
        tarjeta_pad = tarjeta.zfill(8) if tarjeta.isdigit() else tarjeta
        return f"{tarjeta_pad}{val_an}{canal}"
        
    elif canal == "INT-HBK - Multired Virtual (HBK)":
        return f"{tarjeta}{val_an}{hora_de_evento}{canal}"
        
    elif canal == "VRM - VISA (TMGDV)":
        return f"{tarjeta}{val_an}{canal}"
        
    elif canal == "MC - MasterCard (TC)":
        return f"{dni}{val_an}MC - MasterCard (TC)"
        
    elif canal == "FRM - FRM":
        return f"{tarjeta}{canal}"
        
    elif canal == "Batch VISA N7 - Batch VISA N7":
        tarjeta_mask = f"{tarjeta[:6]}******{tarjeta[-4:]}" if len(tarjeta) >= 10 else tarjeta
        return f"{tarjeta_mask}{val_an}{hora_de_evento}{canal}"
        
    elif canal == "BATCH – GIROS - BATCH – GIROS":
        try:
            formatted_date = pd.to_datetime(dia_de_evento).strftime('%Y-%m-%d')
        except:
            formatted_date = str(dia_de_evento)
        return f"{nombre_cliente}{formatted_date}{hora_de_evento}{canal}"
        
    else:
        cuenta_pad = cuenta if len(cuenta) == 11 else f"0{cuenta}"
        return f"{cuenta_pad}{val_an}{canal}"

def main():
    print("Iniciando procesamiento de cruces con reportes y celdas resaltadas...")
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    nocruza_path = os.path.join(script_dir, "NOCRUZA.xlsx")
    basewf_path = os.path.join(script_dir, "BASE WF.xlsx")
    output_corregido_path = os.path.join(script_dir, "devuelto corregido.xlsx")
    output_erroneo_path = os.path.join(script_dir, "registros erroneos.xlsx")
    
    # Load data
    print(f"Leyendo {nocruza_path}...")
    df_no = pd.read_excel(nocruza_path)
    
    print(f"Leyendo {basewf_path} (hoja 'REGISTROS WEBFORM')...")
    df_base = pd.read_excel(basewf_path, sheet_name="REGISTROS WEBFORM")
    
    # Standarize accounts for matching (lstrip '0' to match both '04754518397' and '4754518397')
    df_base['CUENTA_str'] = df_base['CUENTA EMISORA'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True).str.lstrip('0')
    df_no['CUENTA_str'] = df_no['NroCuenta'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True).str.lstrip('0')
    
    erroneous_records = []
    corrected_rows = []
    
    print("\nProcesando y corrigiendo registros...")
    
    for idx, row in df_no.iterrows():
        acc = row['CUENTA_str']
        candidates = df_base[df_base['CUENTA_str'] == acc]
        
        if len(candidates) == 0:
            print(f"Fila {idx:<3} | Cuenta {acc:<8} | ERROR: No se encontró en BASE WF")
            continue
            
        no_date = pd.to_datetime(row['FechaTransaccion']).strftime('%Y-%m-%d')
        no_time = str(row['HoraTransaccion']).strip()
        if len(no_time) == 5:
            no_time += ":00"
        elif len(no_time) > 8:
            no_time = no_time[:8]
            
        no_card = str(row['NroTarjeta']).strip()
        
        # Rank candidates to find the correct one
        best_candidate_idx = None
        best_score = -1
        
        for midx, mrow in candidates.iterrows():
            com = str(mrow['COMUNICACION 1'])
            score = 0
            
            if no_time in com:
                score += 10
            elif no_time[:5] in com:
                score += 5
            if no_date in com:
                score += 5
            formatted_date_slash = pd.to_datetime(no_date).strftime('%d/%m/%Y')
            if formatted_date_slash in com:
                score += 5
            base_time = str(mrow['HORA DE EVENTO']).strip()
            if base_time == no_time[:5]:
                score += 3
            base_date = str(mrow['DIA DE EVENTO']).strip()
            if base_date.startswith(no_date):
                score += 2
                
            if score > best_score:
                best_score = score
                best_candidate_idx = midx
                
        if best_candidate_idx is not None:
            orig_row = df_base.loc[best_candidate_idx].copy()
            corrected_row = orig_row.copy()
            
            com = str(orig_row['COMUNICACION 1'])
            
            # Check for CANAL mismatch
            # Deduce expected canal from rule
            regla_no = str(row.get('Regla', '')).upper()
            expected_canal = orig_row['CANAL']
            if 'BM' in regla_no:
                expected_canal = "BM - Banca Móvil (App)"
            elif 'HBK' in regla_no:
                expected_canal = "INT-HBK - Multired Virtual (HBK)"
            elif 'VRM' in regla_no:
                expected_canal = "VRM - VISA (TMGDV)"
                
            orig_canal = str(orig_row['CANAL']).strip()
            canal_changed = False
            if orig_canal != expected_canal:
                corrected_row['CANAL'] = expected_canal
                canal_changed = True
                print(f"F{idx:<3} | {acc:<10} | CANAL           | {orig_canal:<20} | {expected_canal:<20}")
            
            # 1. Correct CARD
            orig_card = str(orig_row['TARJETA']).strip().replace('.0', '')
            last_4_no = no_card[-4:]
            last_4_orig = orig_card[-4:] if len(orig_card) >= 4 else ""
            
            card_changed = False
            if last_4_no != last_4_orig or orig_card == "0000000000000000" or orig_card == "":
                corrected_row['TARJETA'] = no_card
                card_changed = True
                print(f"F{idx:<3} | {acc:<10} | TARJETA         | {orig_card:<20} | {no_card:<20}")
            
            # 2. Correct TIME (only HH:MM as requested!)
            orig_time = str(orig_row['HORA DE EVENTO']).strip()
            if len(orig_time) > 5:
                orig_time = orig_time[:5]
                
            parsed_time = parse_time_from_text(com)
            correct_time = parsed_time if parsed_time else no_time[:5]
            if len(correct_time) > 5:
                correct_time = correct_time[:5]
                
            time_changed = False
            if orig_time != correct_time or orig_row['HORA DE EVENTO'] == "00:00" or pd.isna(orig_row['HORA DE EVENTO']):
                corrected_row['HORA DE EVENTO'] = correct_time
                time_changed = True
                print(f"F{idx:<3} | {acc:<10} | HORA DE EVENTO  | {orig_time:<20} | {correct_time:<20}")
            else:
                corrected_row['HORA DE EVENTO'] = correct_time
                
            # 3. Correct DATE
            orig_date_raw = str(orig_row['DIA DE EVENTO']).strip()
            try:
                orig_date = pd.to_datetime(orig_date_raw).strftime('%Y-%m-%d')
            except:
                orig_date = orig_date_raw
                
            parsed_date = parse_date_from_text(com)
            correct_date = parsed_date if parsed_date else no_date
            
            date_changed = False
            if orig_date != correct_date:
                corrected_row['DIA DE EVENTO'] = correct_date
                date_changed = True
                print(f"F{idx:<3} | {acc:<10} | DIA DE EVENTO   | {orig_date:<20} | {correct_date:<20}")
                
            # 4. Correct KEY 'L' using the Excel logic
            orig_key = str(orig_row['L']).strip()
            new_key = calculate_key_L(corrected_row)
            corrected_row['L'] = new_key
            if orig_key != new_key:
                print(f"F{idx:<3} | {acc:<10} | Llave 'L'       | {orig_key:<20} | {new_key:<20}")
                
            any_changed = card_changed or time_changed or date_changed or canal_changed
            
            if any_changed:
                # Build OBS
                obs_list = []
                comentarios = []
                
                if date_changed and time_changed:
                    obs_list.append("DIA Y HORA DE EVENTO INCORRECTO")
                    comentarios.append(f"Fecha corregida de '{orig_date}' a '{correct_date}' y Hora de '{orig_time}' a '{correct_time}'")
                elif date_changed:
                    obs_list.append("DIA EVENTO INCORRECTO")
                    comentarios.append(f"Fecha corregida de '{orig_date}' a '{correct_date}'")
                elif time_changed:
                    obs_list.append("HORA DE EVENTO INCORRECTA")
                    comentarios.append(f"Hora corregida de '{orig_time}' a '{correct_time}'")
                    
                if card_changed:
                    obs_list.append("TARJETA INCORRECTA")
                    comentarios.append(f"Tarjeta corregida de '{orig_card}' a '{no_card}'")
                    
                if canal_changed:
                    obs_list.append("CANAL INCORRECTO")
                    comentarios.append(f"Canal corregido de '{orig_canal}' a '{expected_canal}'")
                
                obs_str = " / ".join(obs_list)
                comentario_str = " | ".join(comentarios)
                
                # Save details for reporting
                erroneous_records.append({
                    'orig_row': orig_row,
                    'errors': {
                        'TARJETA': card_changed,
                        'HORA DE EVENTO': time_changed,
                        'DIA DE EVENTO': date_changed,
                        'CANAL': canal_changed
                    },
                    'resumen': {
                        'FECHA DE OPERACION': correct_date,
                        'CANAL': expected_canal,
                        'TJ / CTA / DNI': str(orig_row.get('DNI', '')).replace('.0', ''),
                        'ASESOR ( SEGUN WF)': orig_row.get('GESTOR', ''),
                        'SUPERVISOR': '',
                        'OBS': obs_str,
                        'COMENTARIO ANALISTA': comentario_str
                    }
                })
            else:
                print(f"F{idx:<3} | {acc:<10} | (Correcto)      | Sin cambios en tarjeta/fecha/hora/canal")
                
            # Remove helper column and append
            if 'CUENTA_str' in corrected_row:
                corrected_row = corrected_row.drop('CUENTA_str')
            corrected_rows.append(corrected_row)
            
    # Save Erroneous Records and Resumen
    if len(erroneous_records) > 0:
        # Create DataFrames
        rows_err = []
        rows_res = []
        for r in erroneous_records:
            # Drop helper column
            r_data = r['orig_row'].copy()
            if 'CUENTA_str' in r_data:
                r_data = r_data.drop('CUENTA_str')
            rows_err.append(r_data)
            rows_res.append(r['resumen'])
            
        df_err = pd.DataFrame(rows_err)
        df_res = pd.DataFrame(rows_res)
        
        # Write to registros erroneos.xlsx with highlights
        print(f"\nGuardando {len(df_err)} registros erróneos con resaltado y resumen en {output_erroneo_path}...")
        with pd.ExcelWriter(output_erroneo_path, engine='openpyxl') as writer:
            df_err.to_excel(writer, sheet_name="Registros Erroneos", index=False)
            df_res.to_excel(writer, sheet_name="Resumen Errores", index=False)
            
            # Apply styling/highlights
            workbook = writer.book
            worksheet = workbook["Registros Erroneos"]
            
            # Red fill for incorrect cells
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Get column name to 1-based index mapping
            col_map = {col_name: col_idx + 1 for col_idx, col_name in enumerate(df_err.columns)}
            
            for row_idx, r in enumerate(erroneous_records):
                excel_row = row_idx + 2 # Header is row 1
                for col_name, is_error in r['errors'].items():
                    if is_error and col_name in col_map:
                        col_excel_idx = col_map[col_name]
                        cell = worksheet.cell(row=excel_row, column=col_excel_idx)
                        cell.fill = red_fill
    else:
        print("\nNo se encontraron registros verdaderamente erróneos.")
        # Create empty files
        with pd.ExcelWriter(output_erroneo_path) as writer:
            pd.DataFrame(columns=[c for c in df_base.columns if c != 'CUENTA_str']).to_excel(writer, sheet_name="Registros Erroneos", index=False)
            pd.DataFrame(columns=['FECHA DE OPERACION', 'CANAL', 'TJ / CTA / DNI', 'ASESOR ( SEGUN WF)', 'SUPERVISOR', 'OBS', 'COMENTARIO ANALISTA']).to_excel(writer, sheet_name="Resumen Errores", index=False)
            
    # Save Corrected Records split by TIPO DE GESTIÓN
    df_corrected = pd.DataFrame(corrected_rows)
    
    # Format CUENTA EMISORA to ensure leading zero for 10-digit accounts
    if len(df_corrected) > 0 and 'CUENTA EMISORA' in df_corrected.columns:
        df_corrected['CUENTA EMISORA'] = df_corrected['CUENTA EMISORA'].apply(
            lambda x: f"0{str(x).strip().split('.')[0]}" if pd.notna(x) and str(x).strip().split('.')[0].isdigit() and len(str(x).strip().split('.')[0]) == 10 else x
        )
        
    df_corrected_unique = df_corrected.drop_duplicates()
    
    # Separate OUTBOUND and INBOUND
    df_outbound = df_corrected_unique[df_corrected_unique['TIPO DE GESTIÓN'].astype(str).str.upper().str.contains('OUT')]
    df_inbound = df_corrected_unique[df_corrected_unique['TIPO DE GESTIÓN'].astype(str).str.upper().str.contains('IN')]
    
    print(f"Guardando {len(df_corrected_unique)} registros corregidos (OUTBOUND: {len(df_outbound)}, INBOUND: {len(df_inbound)}) en {output_corregido_path}...")
    with pd.ExcelWriter(output_corregido_path) as writer:
        df_outbound.to_excel(writer, sheet_name="OUTBOUND", index=False)
        df_inbound.to_excel(writer, sheet_name="INBOUND", index=False)
        
    print("¡Proceso completado exitosamente!")

if __name__ == "__main__":
    main()
