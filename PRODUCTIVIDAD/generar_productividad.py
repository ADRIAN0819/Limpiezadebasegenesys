import pandas as pd
import glob
import math
import unicodedata
import re

def format_seconds_to_hhmmss(seconds):
    if pd.isna(seconds):
        return "00:00:00"
    try:
        seconds = float(seconds)
    except (ValueError, TypeError):
        return "00:00:00"
        
    if math.isnan(seconds):
        return "00:00:00"
    
    hh = int(seconds // 3600)
    mm = int((seconds % 3600) // 60)
    ss = int(seconds % 60)
    return f"{hh:02d}:{mm:02d}:{ss:02d}"

def normalize_name(text):
    if not isinstance(text, str):
        return ""
    text = text.upper()
    text = "".join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
    text = text.replace("", "Ñ")
    text = text.replace("NO", "NIÑO")
    text = text.replace("NUEZ", "NUÑEZ")
    text = text.replace("SNCHEZ", "SANCHEZ")
    text = text.replace("DOMNGUEZ", "DOMINGUEZ")
    text = text.replace("PREZ", "PEREZ")
    text = re.sub(r'[^A-ZÑ\s]', ' ', text)
    return " ".join(text.split())

def parse_time_to_seconds(val):
    if pd.isna(val):
        return None
    s = str(val).strip()
    if not s:
        return None
    if hasattr(val, 'hour') and hasattr(val, 'minute'):
        return val.hour * 3600 + val.minute * 60 + getattr(val, 'second', 0)
    match = re.search(r'(\d{1,2}):(\d{2})(?::(\d{2}))?', s)
    if match:
        hh = int(match.group(1))
        mm = int(match.group(2))
        ss = int(match.group(3)) if match.group(3) else 0
        return hh * 3600 + mm * 60 + ss
    return None

def procesar_productividad(cutoff_time=None):
    # 1. Leer el archivo AG
    try:
        df_ag = pd.read_excel('AG.xlsx', header=None)
    except Exception as e:
        print(f"Error al leer AG.xlsx: {e}")
        return
    
    # Extraer la lista de AGs válidos de la primera columna (índice 0)
    lista_ag = df_ag[0].dropna().astype(str).tolist()
    lista_ag = [ag.strip() for ag in lista_ag if str(ag).strip()]
    
    # Preparar df_ag para hacer el cruce (y obtener Turno de la 3era columna si existe)
    if len(df_ag.columns) >= 3:
        df_ag_clean = df_ag[[0, 1, 2]].dropna(subset=[0]).rename(columns={0: 'AG', 1: 'Nombre', 2: 'Turno'})
    else:
        df_ag_clean = df_ag[[0, 1]].dropna(subset=[0]).rename(columns={0: 'AG', 1: 'Nombre'})
        df_ag_clean['Turno'] = ""

    df_ag_clean['AG'] = df_ag_clean['AG'].astype(str).str.strip()
    df_ag_clean['Nombre_norm'] = df_ag_clean['Nombre'].apply(normalize_name)
    df_ag_clean['Turno'] = df_ag_clean['Turno'].fillna("").astype(str).str.strip()
    
    # 2. Encontrar el archivo de resumen
    archivos_resumen = glob.glob('*Resumen del rendimiento de agente.csv')
    if not archivos_resumen:
        print("No se encontró ningún archivo que termine en 'Resumen del rendimiento de agente.csv'")
        return
        
    archivo_resumen = archivos_resumen[0]
    print(f"Procesando archivo: {archivo_resumen}")
    
    # Leer el csv (está separado por punto y coma usualmente)
    try:
        df_resumen = pd.read_csv(archivo_resumen, sep=';', encoding='utf-8')
    except UnicodeDecodeError:
        df_resumen = pd.read_csv(archivo_resumen, sep=';', encoding='latin1')
    
    # Validar que las columnas existan
    columnas_necesarias = ['Nombre del agente', 'Contestadas', 'Saliente', 'Manejo medio']
    for col in columnas_necesarias:
        if col not in df_resumen.columns:
            print(f"La columna '{col}' no se encontró en el archivo de resumen.")
            return
            
    # Función para extraer el AG del nombre del agente
    def extract_ag(nombre):
        nombre_str = str(nombre)
        for ag in lista_ag:
            if ag in nombre_str:
                return ag
        return None
        
    # Extraer el AG que coincida
    df_resumen = df_resumen.copy()
    df_resumen['AG_matched'] = df_resumen['Nombre del agente'].apply(extract_ag)
    df_filtrado = df_resumen[df_resumen['AG_matched'].notnull()].copy()
    
    if df_filtrado.empty:
        print("No se encontraron coincidencias de AG en el archivo de resumen.")
        return
        
    # Unir los datos
    df_resultado = pd.merge(df_filtrado, df_ag_clean[['AG', 'Nombre', 'Turno']], left_on='AG_matched', right_on='AG', how='left')
    df_resultado['Turno'] = df_resultado['Turno'].fillna("")
    
    # Llenar NaNs de métricas con 0 y convertir a entero
    df_resultado['Contestadas'] = df_resultado['Contestadas'].fillna(0).astype(int)
    df_resultado['Saliente'] = df_resultado['Saliente'].fillna(0).astype(int)
    
    # Formatear el Manejo medio a hh:mm:ss
    df_resultado['Manejo Medio'] = df_resultado['Manejo medio'].apply(format_seconds_to_hhmmss)
    
    # 3. Procesar los archivos de Webform (INBOUND y OUTBOUND) para contar registros
    inbound_files = glob.glob("*INBOUND*.xlsx")
    outbound_files = glob.glob("*OUTBOUND*.xlsx")
    
    # Diccionarios para almacenar la cuenta de registros por AG
    in_counts = {ag: 0 for ag in lista_ag}
    out_counts = {ag: 0 for ag in lista_ag}
    
    # Caché de coincidencia para evitar recalcular
    match_cache = {}
    
    cutoff_sec = parse_time_to_seconds(cutoff_time) if cutoff_time else None

    def find_ag_for_gestor(gestor_name):
        if not isinstance(gestor_name, str) or not gestor_name.strip():
            return None
        gestor_clean = gestor_name.strip()
        if gestor_clean in match_cache:
            return match_cache[gestor_clean]
            
        g_norm = normalize_name(gestor_clean)
        g_tokens = set(g_norm.split())
        if not g_tokens:
            return None
            
        best_ag = None
        best_score = 0
        best_row = None
        
        for idx, row in df_ag_clean.iterrows():
            ag_tokens = set(row['Nombre_norm'].split())
            overlap = g_tokens.intersection(ag_tokens)
            score = len(overlap)
            if score > best_score:
                best_score = score
                best_row = row
            elif score == best_score and best_score > 0:
                cur_diff = abs(len(g_tokens) - len(set(best_row['Nombre_norm'].split())))
                new_diff = abs(len(g_tokens) - len(ag_tokens))
                if new_diff < cur_diff:
                    best_row = row
                    
        if best_row is not None and best_score >= 2:
            resolved_ag = best_row['AG']
        else:
            resolved_ag = None
            
        match_cache[gestor_clean] = resolved_ag
        return resolved_ag

    # Procesar INBOUND
    if inbound_files:
        archivo_in = inbound_files[0]
        print(f"Procesando Webform INBOUND: {archivo_in}")
        try:
            df_in = pd.read_excel(archivo_in)
            if 'GESTOR' in df_in.columns:
                if cutoff_sec is not None and 'HORA' in df_in.columns:
                    df_in = df_in[df_in['HORA'].apply(lambda x: (s := parse_time_to_seconds(x)) is not None and s <= cutoff_sec)]
                for gestor in df_in['GESTOR'].dropna():
                    ag_code = find_ag_for_gestor(gestor)
                    if ag_code:
                        in_counts[ag_code] += 1
            else:
                print("Advertencia: No se encontró la columna 'GESTOR' en el archivo INBOUND.")
        except Exception as e:
            print(f"Error al procesar archivo INBOUND: {e}")
    else:
        print("No se encontró ningún archivo de Webform INBOUND (*INBOUND*.xlsx)")

    # Procesar OUTBOUND
    if outbound_files:
        archivo_out = outbound_files[0]
        print(f"Procesando Webform OUTBOUND: {archivo_out}")
        try:
            df_out = pd.read_excel(archivo_out)
            if 'GESTOR' in df_out.columns:
                if cutoff_sec is not None and 'HORA' in df_out.columns:
                    df_out = df_out[df_out['HORA'].apply(lambda x: (s := parse_time_to_seconds(x)) is not None and s <= cutoff_sec)]
                for gestor in df_out['GESTOR'].dropna():
                    ag_code = find_ag_for_gestor(gestor)
                    if ag_code:
                        out_counts[ag_code] += 1
            else:
                print("Advertencia: No se encontró la columna 'GESTOR' en el archivo OUTBOUND.")
        except Exception as e:
            print(f"Error al procesar archivo OUTBOUND: {e}")
    else:
        print("No se encontró ningún archivo de Webform OUTBOUND (*OUTBOUND*.xlsx)")

    # Agregar las columnas individuales WF IN y WF OUT
    df_resultado['WF IN'] = df_resultado['AG'].apply(lambda ag: in_counts.get(ag, 0))
    df_resultado['WF OUT'] = df_resultado['AG'].apply(lambda ag: out_counts.get(ag, 0))

    # Quitar agentes sin actividad en llamadas y ordenar por salientes
    df_resultado = df_resultado[
        ~((df_resultado['Contestadas'] == 0) & (df_resultado['Saliente'] == 0))
    ].copy()
    df_resultado = df_resultado.sort_values(by='Saliente', ascending=False)
    
    # Seleccionar y ordenar las columnas finales
    df_resultado = df_resultado[['AG', 'Nombre', 'Turno', 'Saliente', 'Contestadas', 'Manejo Medio', 'WF IN', 'WF OUT']]
    
    # Exportar a Excel con diseño premium usando openpyxl
    nombre_salida = 'Productividad_Agentes.xlsx'
    
    try:
        with pd.ExcelWriter(nombre_salida, engine='openpyxl') as writer:
            df_resultado.to_excel(writer, sheet_name='Productividad Agentes', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Productividad Agentes']
            
            # Asegurar visualización de líneas de cuadrícula
            worksheet.views.sheetView[0].showGridLines = True
            
            # Importar librerías de estilos de openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Paleta de colores Premium
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Azul Navy
            header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
            
            # Zebra striping
            zebra_fill = PatternFill(start_color='F2F5F8', end_color='F2F5F8', fill_type='solid') # Gris/Azul suave
            
            # Bordes finos
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )
            
            # Fuentes y alineaciones
            font_regular = Font(name='Segoe UI', size=10)
            align_center = Alignment(horizontal='center', vertical='center')
            align_left = Alignment(horizontal='left', vertical='center')
            
            # Aplicar estilos a las cabeceras (Fila 1)
            for col_num in range(1, len(df_resultado.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
                cell.border = thin_border
                
            # Aplicar estilos a las celdas de datos
            for row_idx in range(2, worksheet.max_row + 1):
                is_even = (row_idx % 2 == 0)
                for col_idx in range(1, len(df_resultado.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = font_regular
                    cell.border = thin_border
                    
                    # Zebra striping
                    if is_even:
                        cell.fill = zebra_fill
                        
                    # Alineaciones según columna
                    col_name = df_resultado.columns[col_idx - 1]
                    if col_name == 'Nombre':
                        cell.alignment = align_left
                    else:
                        cell.alignment = align_center
            
            # Auto-ajustar ancho de las columnas
            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
        print(f"Proceso completado exitosamente. Archivo generado con diseño: {nombre_salida}")
    except Exception as e:
        print(f"Error al escribir y dar diseño al archivo Excel: {e}")

if __name__ == '__main__':
    procesar_productividad()
